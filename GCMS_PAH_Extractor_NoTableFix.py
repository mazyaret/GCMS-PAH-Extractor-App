#!/usr/bin/env python3
"""
GC/MS PAH Extractor

Purpose:
- Reads the standard GC/MS Excel output where row 1 contains analyte result blocks
  and row 2 contains fields such as RT, Final Conc., Area, Ratio.
- Extracts the "Final Conc." column for each target PAH/surrogate.
- Applies the dilution factor (DF), either parsed from sample names like "8X"
  or provided in a DF template workbook.
- Exports a clean Excel workbook with raw concentration, DF-corrected concentration,
  and surrogate recovery %.

Usage examples:
    python GCMS_PAH_Extractor.py "PAH batch 04292026.xlsx"
    python GCMS_PAH_Extractor.py "PAH batch 04292026.xlsx" --df-template "GCMS_DF_Template.xlsx"
    python GCMS_PAH_Extractor.py "PAH batch 04292026.xlsx" --output "processed.xlsx"

Dependency:
    pip install openpyxl
"""

from __future__ import annotations

import argparse
import datetime
import re
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule
from openpyxl.comments import Comment

ANALYTES: List[Tuple[str, str, Optional[float]]] = [
    ("Nitrobenzene-d5", "Surrogate", 0.5),
    ("Naphthalene", "Target", None),
    ("2-fluorobiphenyl", "Surrogate", 0.5),
    ("Acenaphthylene", "Target", None),
    ("Acenaphthene", "Target", None),
    ("Fluorene-d10", "Surrogate", 0.5),
    ("Fluorene", "Target", None),
    ("2,4,6 Tribromophenol", "Surrogate", 0.5),
    ("Phenanthrene", "Target", None),
    ("Anthracene", "Target", None),
    ("Fluoranthene", "Target", None),
    ("Pyrene-d10", "Surrogate", 0.5),
    ("Pyrene", "Target", None),
    ("P-terphenyl-d14", "Surrogate", 0.5),
    ("Benz[a]anthracene", "Target", None),
    ("Chrysene", "Target", None),
    ("Benzo[b]fluoranthene", "Target", None),
    ("Benzo[k]fluoranthene", "Target", None),
    ("Benzo[a]pyrene-d12", "Surrogate", 0.5),
    ("Benzo[a]pyrene", "Target", None),
    ("Indeno[1,2,3-cd]pyrene", "Target", None),
    ("Dibenz[a,h]anthracene", "Target", None),
    ("Benzo[g,h,i]perylene", "Target", None),
]

IGNORE_PATTERNS_DEFAULT = [
    r"(?i)\bug/?ml\b",
    r"(?i)\bugml\b",
    r"(?i)surrog[_\s-]*mrm",
    r"(?i)\bstandard\b",
    r"(?i)\bstd\b",
    r"(?i)\bcal\b",
    r"(?i)\bblank\b",
]


def norm_name(value: Any) -> str:
    text = str(value or "").strip()
    text = re.sub(r"\s+", " ", text)
    return text.replace("–", "-").replace("—", "-").lower()


def parse_df(sample_name: str) -> Tuple[float, str]:
    text = str(sample_name or "")
    matches = re.findall(r"(?i)(?<![A-Za-z0-9])(\d+(?:\.\d+)?)\s*[x×]\b", text)
    if matches:
        try:
            return float(matches[-1]), "Parsed from sample name"
        except ValueError:
            pass
    return 1.0, "Default 1"


def should_ignore_sample(sample_name: str, ignore_patterns: List[str] = IGNORE_PATTERNS_DEFAULT) -> bool:
    text = str(sample_name or "")
    if not text.strip():
        return True
    return any(re.search(pattern, text) for pattern in ignore_patterns)


def load_df_overrides(df_template_path: Optional[Path]) -> Dict[str, float]:
    if not df_template_path:
        return {}
    wb = load_workbook(df_template_path, data_only=True)
    if "Dilution_Factors" in wb.sheetnames:
        ws = wb["Dilution_Factors"]
    else:
        ws = wb[wb.sheetnames[0]]

    headers = {norm_name(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)}
    sample_col = headers.get("sample name") or headers.get("sample")
    df_col = headers.get("dilution factor") or headers.get("df")
    if not sample_col or not df_col:
        raise ValueError("DF template must contain headers 'Sample Name' and 'Dilution Factor'.")

    overrides: Dict[str, float] = {}
    for row in range(2, ws.max_row + 1):
        sample = ws.cell(row, sample_col).value
        df = ws.cell(row, df_col).value
        if sample in (None, "") or df in (None, ""):
            continue
        try:
            df_float = float(df)
        except (TypeError, ValueError):
            continue
        overrides[str(sample)] = df_float
        overrides[norm_name(sample)] = df_float
    return overrides


def find_header_rows(ws) -> Tuple[int, int]:
    for r in range(1, min(ws.max_row, 25) + 1):
        values = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 60) + 1)]
        if "Name" in values and "Final Conc." in values:
            return r - 1, r
    return 1, 2


def find_column(ws, subheader_row: int, target: str) -> Optional[int]:
    for c in range(1, ws.max_column + 1):
        if norm_name(ws.cell(subheader_row, c).value) == norm_name(target):
            return c
    return None


def find_acq_col(ws, subheader_row: int) -> Optional[int]:
    for c in range(1, ws.max_column + 1):
        label = norm_name(ws.cell(subheader_row, c).value)
        if "acq" in label and "date" in label:
            return c
    return None


def find_analyte_final_conc_cols(ws, analytes, analyte_header_row: int, subheader_row: int):
    header_analytes = {}
    for c in range(1, ws.max_column + 1):
        value = ws.cell(analyte_header_row, c).value
        if isinstance(value, str) and value.strip().lower().endswith(" results"):
            analyte = re.sub(r"\s+Results\s*$", "", value.strip(), flags=re.I)
            header_analytes[norm_name(analyte)] = c

    result = {}
    missing = []
    for analyte, _, _ in analytes:
        start_col = header_analytes.get(norm_name(analyte))
        if not start_col:
            missing.append(analyte)
            continue
        final_col = None
        for cc in range(start_col, min(ws.max_column, start_col + 10) + 1):
            if norm_name(ws.cell(subheader_row, cc).value) == "final conc.":
                final_col = cc
                break
        if final_col is None:
            missing.append(analyte)
        else:
            result[analyte] = final_col
    return result, missing


def extract_gcms(input_xlsx: Path, df_overrides=None, keep_ignored=False):
    wb_in = load_workbook(input_xlsx, data_only=True, read_only=False)
    ws = wb_in[wb_in.sheetnames[0]]

    analyte_header_row, subheader_row = find_header_rows(ws)
    name_col = find_column(ws, subheader_row, "Name") or 3
    datafile_col = find_column(ws, subheader_row, "Data File")
    acq_col = find_acq_col(ws, subheader_row)
    analyte_cols, missing = find_analyte_final_conc_cols(ws, ANALYTES, analyte_header_row, subheader_row)

    samples = []
    for row_idx in range(subheader_row + 1, ws.max_row + 1):
        name = ws.cell(row_idx, name_col).value
        if name is None or str(name).strip() == "":
            continue

        ignored = should_ignore_sample(str(name))
        if ignored and not keep_ignored:
            continue

        df, df_source = parse_df(str(name))
        if df_overrides:
            if str(name) in df_overrides:
                df = df_overrides[str(name)]
                df_source = "Manual override"
            elif norm_name(name) in df_overrides:
                df = df_overrides[norm_name(name)]
                df_source = "Manual override"

        record = {
            "row": row_idx,
            "sample_name": str(name),
            "data_file": ws.cell(row_idx, datafile_col).value if datafile_col else None,
            "acq_datetime": ws.cell(row_idx, acq_col).value if acq_col else None,
            "dilution_factor": float(df) if df not in (None, "") else 1.0,
            "df_source": df_source,
            "ignored": ignored,
            "values": {},
        }

        for analyte, _, _ in ANALYTES:
            col = analyte_cols.get(analyte)
            value = ws.cell(row_idx, col).value if col else None
            if isinstance(value, str):
                try:
                    value = float(value)
                except ValueError:
                    pass
            record["values"][analyte] = value
        samples.append(record)

    meta = {
        "input_file": str(input_xlsx),
        "sheet_name": ws.title,
        "analyte_header_row": analyte_header_row,
        "subheader_row": subheader_row,
        "name_col": get_column_letter(name_col),
        "data_file_col": get_column_letter(datafile_col) if datafile_col else None,
        "acq_col": get_column_letter(acq_col) if acq_col else None,
        "analyte_final_conc_cols": {k: get_column_letter(v) for k, v in analyte_cols.items()},
        "missing_analytes": missing,
        "extracted_samples": len(samples),
        "ignored_patterns": IGNORE_PATTERNS_DEFAULT,
    }
    return samples, meta


def add_table(ws, table_name: str, start_row: int, start_col: int, end_row: int, end_col: int) -> None:
    """Disabled intentionally to avoid Excel repair logs for table XML parts."""
    return


def write_processed_workbook(samples, meta, output_path: Path) -> None:
    wb = Workbook()
    ws_info = wb.active
    ws_info.title = "Instructions"
    ws_samples = wb.create_sheet("Sample_Info")
    ws_wide = wb.create_sheet("Wide_Data")
    ws_long = wb.create_sheet("Long_Data")
    ws_qc = wb.create_sheet("Recovery_QC")
    ws_config = wb.create_sheet("Config_Analytes")
    ws_run = wb.create_sheet("Run_Info")

    title_fill = PatternFill("solid", fgColor="DDEBF7")
    header_fill = PatternFill("solid", fgColor="E2F0D9")
    surrogate_fill = PatternFill("solid", fgColor="FFF2CC")
    target_fill = PatternFill("solid", fgColor="F2F2F2")
    border = Border(bottom=Side(style="thin", color="D9E2F3"))
    dark_font = Font(name="Aptos", bold=True, size=11, color="1F4E78")

    instructions = [
        ["GC/MS PAH Extraction Platform", ""],
        ["Purpose", "Extracts each analyte's Final Conc. from the original GC/MS sheet, applies dilution factors, and calculates surrogate recovery %."],
        ["DF formula", "Corrected concentration = Raw Final Conc. × Dilution Factor."],
        ["Recovery formula", "Recovery % = Corrected surrogate concentration ÷ surrogate spike concentration. Default spike concentration is 0.5 µg/mL."],
        ["Ignored rows", "By default the tool ignores calibration/standard-like names containing ugml, ug/ml, surrog_mrm, standard, std, cal, or blank."],
        ["Main sheets", "Sample_Info = sample names and DFs; Wide_Data = one row per sample; Long_Data = tidy table; Recovery_QC = surrogate recoveries; Config_Analytes = analyte list; Run_Info = extraction metadata."],
    ]
    for r, row in enumerate(instructions, 1):
        for c, value in enumerate(row, 1):
            ws_info.cell(r, c).value = value
    ws_info["A1"].font = Font(name="Aptos Display", bold=True, size=18, color="1F4E78")
    ws_info["A1"].fill = title_fill
    ws_info.merge_cells("A1:B1")
    for r in range(2, len(instructions) + 1):
        ws_info.cell(r, 1).font = dark_font
        ws_info.cell(r, 1).fill = header_fill
        ws_info.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
        ws_info.row_dimensions[r].height = 34
    ws_info.column_dimensions["A"].width = 24
    ws_info.column_dimensions["B"].width = 110

    sample_headers = ["Sample Name", "Data File", "Acq. Date-Time", "Dilution Factor", "DF Source", "Notes"]
    ws_samples.append(sample_headers)
    for rec in samples:
        ws_samples.append([rec["sample_name"], rec["data_file"], rec["acq_datetime"], rec["dilution_factor"], rec["df_source"], ""])
    for c in range(1, len(sample_headers) + 1):
        ws_samples.cell(1, c).font = Font(name="Aptos", bold=True, color="1F4E78")
        ws_samples.cell(1, c).fill = header_fill
    ws_samples.freeze_panes = "A2"
    ws_samples.auto_filter.ref = f"A1:{get_column_letter(len(sample_headers))}{max(1, len(samples) + 1)}"
    add_table(ws_samples, "SampleInfoTable", 1, 1, len(samples) + 1, len(sample_headers))
    for row in range(2, len(samples) + 2):
        ws_samples.cell(row, 3).number_format = "yyyy-mm-dd hh:mm"
        ws_samples.cell(row, 4).number_format = "0.###"
    for i, width in enumerate([28, 28, 22, 16, 22, 30], 1):
        ws_samples.column_dimensions[get_column_letter(i)].width = width
    ws_samples["D1"].comment = Comment("The tool parses values such as 8X from sample names. Use the DF template for manual overrides.", "ChatGPT")

    config_headers = ["Analyte", "Role", "Surrogate Spike Conc. (µg/mL)", "Include in Output", "QC Low Recovery", "QC High Recovery"]
    ws_config.append(config_headers)
    for analyte, role, spike in ANALYTES:
        ws_config.append([analyte, role, spike if spike is not None else "", "Yes", 0.70 if role == "Surrogate" else "", 1.30 if role == "Surrogate" else ""])
    for c in range(1, len(config_headers) + 1):
        ws_config.cell(1, c).font = Font(name="Aptos", bold=True, color="1F4E78")
        ws_config.cell(1, c).fill = header_fill
    for r in range(2, len(ANALYTES) + 2):
        fill = surrogate_fill if ws_config.cell(r, 2).value == "Surrogate" else target_fill
        for c in range(1, len(config_headers) + 1):
            ws_config.cell(r, c).fill = fill
            ws_config.cell(r, c).border = border
        ws_config.cell(r, 5).number_format = "0%"
        ws_config.cell(r, 6).number_format = "0%"
    add_table(ws_config, "ConfigAnalytesTable", 1, 1, len(ANALYTES) + 1, len(config_headers))
    ws_config.freeze_panes = "A2"
    for i, width in enumerate([28, 14, 26, 18, 18, 19], 1):
        ws_config.column_dimensions[get_column_letter(i)].width = width
    ws_config["C1"].comment = Comment("Default is 0.5 µg/mL. Edit this in the script if your surrogate spike concentration changes.", "ChatGPT")

    wide_headers = ["Sample Name", "Data File", "Acq. Date-Time", "Dilution Factor"]
    for analyte, role, _ in ANALYTES:
        wide_headers.append(f"{analyte} Raw Final Conc.")
        wide_headers.append(f"{analyte} Corrected Conc.")
        if role == "Surrogate":
            wide_headers.append(f"{analyte} Recovery %")
    ws_wide.append(wide_headers)
    for rec in samples:
        row = [rec["sample_name"], rec["data_file"], rec["acq_datetime"], rec["dilution_factor"]]
        df = rec["dilution_factor"]
        for analyte, role, spike in ANALYTES:
            raw = rec["values"].get(analyte)
            corrected = raw * df if isinstance(raw, (int, float)) else None
            row.extend([raw, corrected])
            if role == "Surrogate":
                recovery = corrected / spike if isinstance(corrected, (int, float)) and spike else None
                row.append(recovery)
        ws_wide.append(row)
    for c in range(1, len(wide_headers) + 1):
        ws_wide.cell(1, c).font = Font(name="Aptos", bold=True, color="1F4E78")
        ws_wide.cell(1, c).fill = header_fill
        ws_wide.cell(1, c).alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    for r in range(2, len(samples) + 2):
        ws_wide.cell(r, 3).number_format = "yyyy-mm-dd hh:mm"
        ws_wide.cell(r, 4).number_format = "0.###"
        for c in range(5, len(wide_headers) + 1):
            header = str(ws_wide.cell(1, c).value)
            ws_wide.cell(r, c).number_format = "0.0%" if "Recovery %" in header else "0.000000"
    ws_wide.freeze_panes = "E2"
    ws_wide.auto_filter.ref = f"A1:{get_column_letter(len(wide_headers))}{len(samples) + 1}"
    add_table(ws_wide, "WideDataTable", 1, 1, len(samples) + 1, len(wide_headers))
    for c in range(1, len(wide_headers) + 1):
        ws_wide.column_dimensions[get_column_letter(c)].width = 26 if c == 1 else (22 if c in (2, 3) else (14 if c == 4 else 15))

    long_headers = ["Sample Name", "Data File", "Acq. Date-Time", "Dilution Factor", "Analyte", "Role", "Raw Final Conc.", "Corrected Conc.", "Surrogate Spike Conc. (µg/mL)", "Recovery %"]
    ws_long.append(long_headers)
    for rec in samples:
        df = rec["dilution_factor"]
        for analyte, role, spike in ANALYTES:
            raw = rec["values"].get(analyte)
            corrected = raw * df if isinstance(raw, (int, float)) else None
            recovery = corrected / spike if role == "Surrogate" and isinstance(corrected, (int, float)) and spike else None
            ws_long.append([rec["sample_name"], rec["data_file"], rec["acq_datetime"], df, analyte, role, raw, corrected, spike if role == "Surrogate" else None, recovery])
    for c in range(1, len(long_headers) + 1):
        ws_long.cell(1, c).font = Font(name="Aptos", bold=True, color="1F4E78")
        ws_long.cell(1, c).fill = header_fill
    for r in range(2, ws_long.max_row + 1):
        ws_long.cell(r, 3).number_format = "yyyy-mm-dd hh:mm"
        ws_long.cell(r, 4).number_format = "0.###"
        ws_long.cell(r, 7).number_format = "0.000000"
        ws_long.cell(r, 8).number_format = "0.000000"
        ws_long.cell(r, 9).number_format = "0.###"
        ws_long.cell(r, 10).number_format = "0.0%"
    ws_long.freeze_panes = "A2"
    ws_long.auto_filter.ref = f"A1:{get_column_letter(len(long_headers))}{ws_long.max_row}"
    add_table(ws_long, "LongDataTable", 1, 1, ws_long.max_row, len(long_headers))
    for i, width in enumerate([26, 22, 22, 14, 28, 12, 16, 16, 24, 14], 1):
        ws_long.column_dimensions[get_column_letter(i)].width = width

    qc_headers = ["Sample Name", "Dilution Factor"]
    surrogates = [(a, s) for a, r, s in ANALYTES if r == "Surrogate"]
    for analyte, _ in surrogates:
        qc_headers.append(f"{analyte} Recovery %")
        qc_headers.append(f"{analyte} QC")
    ws_qc.append(qc_headers)
    low, high = 0.70, 1.30
    for rec in samples:
        row = [rec["sample_name"], rec["dilution_factor"]]
        df = rec["dilution_factor"]
        for analyte, spike in surrogates:
            raw = rec["values"].get(analyte)
            corrected = raw * df if isinstance(raw, (int, float)) else None
            recovery = corrected / spike if isinstance(corrected, (int, float)) and spike else None
            qc = "Check" if recovery is not None and (recovery < low or recovery > high) else ("OK" if recovery is not None else "Missing")
            row.extend([recovery, qc])
        ws_qc.append(row)
    for c in range(1, len(qc_headers) + 1):
        ws_qc.cell(1, c).font = Font(name="Aptos", bold=True, color="1F4E78")
        ws_qc.cell(1, c).fill = header_fill
        ws_qc.cell(1, c).alignment = Alignment(wrap_text=True, horizontal="center")
    for r in range(2, ws_qc.max_row + 1):
        ws_qc.cell(r, 2).number_format = "0.###"
        for c in range(3, len(qc_headers) + 1):
            if "Recovery %" in str(ws_qc.cell(1, c).value):
                ws_qc.cell(r, c).number_format = "0.0%"
    ws_qc.freeze_panes = "C2"
    ws_qc.auto_filter.ref = f"A1:{get_column_letter(len(qc_headers))}{ws_qc.max_row}"
    add_table(ws_qc, "RecoveryQCTable", 1, 1, ws_qc.max_row, len(qc_headers))
    ws_qc.conditional_formatting.add(
        f"A2:{get_column_letter(len(qc_headers))}{ws_qc.max_row}",
        CellIsRule(operator="equal", formula=['"Check"'], fill=PatternFill("solid", fgColor="F4CCCC")),
    )
    for c in range(1, len(qc_headers) + 1):
        ws_qc.column_dimensions[get_column_letter(c)].width = 24 if c <= 2 else 18

    run_rows = [
        ["Created", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Input file", meta.get("input_file")],
        ["Input sheet", meta.get("sheet_name")],
        ["Sample name column", meta.get("name_col")],
        ["Data file column", meta.get("data_file_col")],
        ["Acquisition date column", meta.get("acq_col")],
        ["Analyte header row", meta.get("analyte_header_row")],
        ["Subheader row", meta.get("subheader_row")],
        ["Extracted samples", meta.get("extracted_samples")],
        ["Missing analytes", ", ".join(meta.get("missing_analytes", [])) if meta.get("missing_analytes") else "None"],
        ["Ignored row patterns", "; ".join(meta.get("ignored_patterns", []))],
    ]
    for row in run_rows:
        ws_run.append(row)
    for r in range(1, len(run_rows) + 1):
        ws_run.cell(r, 1).font = dark_font
        ws_run.cell(r, 1).fill = header_fill
        ws_run.cell(r, 2).alignment = Alignment(wrap_text=True)
    ws_run.column_dimensions["A"].width = 28
    ws_run.column_dimensions["B"].width = 100
    ws_run.append([])
    ws_run.append(["Analyte", "Final Conc. Column"])
    for analyte, col_letter in meta.get("analyte_final_conc_cols", {}).items():
        ws_run.append([analyte, col_letter])
    for c in range(1, 3):
        ws_run.cell(13, c).font = Font(name="Aptos", bold=True, color="1F4E78")
        ws_run.cell(13, c).fill = header_fill
    add_table(ws_run, "RunInfoColsTable", 13, 1, ws_run.max_row, 2)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.row_dimensions[1].height = 32
        for row in ws.iter_rows():
            for cell in row:
                cell.font = cell.font.copy(name="Aptos") if cell.font else Font(name="Aptos", size=10)
                cell.alignment = Alignment(vertical="center", wrap_text=cell.alignment.wrap_text)

    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="Extract PAH Final Conc. values from standard GC/MS Excel output.")
    parser.add_argument("input", nargs="?", help="Original GC/MS Excel workbook (.xlsx).")
    parser.add_argument("--df-template", help="Optional Excel file with 'Dilution_Factors' sheet and columns 'Sample Name' and 'Dilution Factor'.")
    parser.add_argument("--output", help="Output processed workbook path.")
    parser.add_argument("--keep-ignored", action="store_true", help="Keep standard/calibration-like rows instead of ignoring them.")
    args = parser.parse_args()

    input_path = Path(args.input) if args.input else None
    if input_path is None:
        try:
            import tkinter as tk
            from tkinter import filedialog
            root = tk.Tk()
            root.withdraw()
            selected = filedialog.askopenfilename(title="Select original GC/MS Excel workbook", filetypes=[("Excel files", "*.xlsx *.xlsm")])
            if selected:
                input_path = Path(selected)
        except Exception:
            pass
    if input_path is None or not input_path.exists():
        raise SystemExit("Please provide a valid input Excel workbook.")

    output_path = Path(args.output) if args.output else input_path.with_name(input_path.stem + "_processed.xlsx")
    df_overrides = load_df_overrides(Path(args.df_template)) if args.df_template else {}

    samples, meta = extract_gcms(input_path, df_overrides=df_overrides, keep_ignored=args.keep_ignored)
    write_processed_workbook(samples, meta, output_path)
    print(f"Done. Extracted {len(samples)} samples.")
    if meta.get("missing_analytes"):
        print("Warning: missing analytes:", ", ".join(meta["missing_analytes"]))
    print(f"Output: {output_path}")


if __name__ == "__main__":
    main()
