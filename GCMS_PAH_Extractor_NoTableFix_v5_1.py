#!/usr/bin/env python3
"""
GC/MS PAH Extractor - NoTableFix v5

Updates through v5:
- Keeps Excel table objects disabled to avoid Excel repair warnings.
- Accepts lab sample submission form columns for DF, fabric size, fabric mass, fuel mass loss, and extract/sample volume.
- Computes normalized values for the 16 priority PAHs:
  surface load (ug/100 cm^2), ug/g fabric, and ug/g fuel mass loss.
- Computes Total PAHs, LMW PAHs, HMW PAHs, and TEQ summaries.
"""

from __future__ import annotations

import argparse
import datetime
import math
import re
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

ANALYTES: List[Tuple[str, str, Optional[float]]] = [
    ("Nitrobenzene-d5", "Surrogate", 0.5),
    ("Naphthalene", "Target", None),
    ("2-fluorobiphenyl", "Surrogate", 0.5),
    ("Acenaphthylene", "Target", None),
    ("Acenaphthene", "Target", None),
    ("Fluorene-d10", "Surrogate", 0.5),
    ("Fluorene", "Target", None),
    ("2,4,6 Tribromophenol", "Surrogate", 0.5),
    ("Phenanthrene", "Target", None),
    ("Anthracene", "Target", None),
    ("Fluoranthene", "Target", None),
    ("Pyrene-d10", "Surrogate", 0.5),
    ("Pyrene", "Target", None),
    ("P-terphenyl-d14", "Surrogate", 0.5),
    ("Benz[a]anthracene", "Target", None),
    ("Chrysene", "Target", None),
    ("Benzo[b]fluoranthene", "Target", None),
    ("Benzo[k]fluoranthene", "Target", None),
    ("Benzo[a]pyrene-d12", "Surrogate", 0.5),
    ("Benzo[a]pyrene", "Target", None),
    ("Indeno[1,2,3-cd]pyrene", "Target", None),
    ("Dibenz[a,h]anthracene", "Target", None),
    ("Benzo[g,h,i]perylene", "Target", None),
]

TARGET_16_PAHS = [
    "Naphthalene",
    "Acenaphthylene",
    "Acenaphthene",
    "Fluorene",
    "Phenanthrene",
    "Anthracene",
    "Fluoranthene",
    "Pyrene",
    "Benz[a]anthracene",
    "Chrysene",
    "Benzo[b]fluoranthene",
    "Benzo[k]fluoranthene",
    "Benzo[a]pyrene",
    "Indeno[1,2,3-cd]pyrene",
    "Dibenz[a,h]anthracene",
    "Benzo[g,h,i]perylene",
]

LMW_PAHS = ["Naphthalene", "Acenaphthylene", "Acenaphthene", "Fluorene", "Phenanthrene", "Anthracene"]
HMW_PAHS = [a for a in TARGET_16_PAHS if a not in LMW_PAHS]

TEF_FACTORS = {
    "Naphthalene": 0.001,
    "Acenaphthylene": 0.001,
    "Acenaphthene": 0.001,
    "Fluorene": 0.001,
    "Phenanthrene": 0.001,
    "Anthracene": 0.01,
    "Fluoranthene": 0.001,
    "Pyrene": 0.001,
    "Benz[a]anthracene": 0.1,
    "Chrysene": 0.01,
    "Benzo[b]fluoranthene": 0.1,
    "Benzo[k]fluoranthene": 0.1,
    "Benzo[a]pyrene": 1,
    "Indeno[1,2,3-cd]pyrene": 0.1,
    "Dibenz[a,h]anthracene": 1,
    "Benzo[g,h,i]perylene": 0.01,
}

# Recovery-correction mapping for the ISU/all-in-one GC/MS format.
# These are extraction surrogates used to estimate recovery for PAHs.
RECOVERY_SURROGATE_MAP_ISU = {
    "Naphthalene": "2-fluorobiphenyl",
    "Acenaphthylene": "2-fluorobiphenyl",
    "Acenaphthene": "2-fluorobiphenyl",
    "Fluorene": "Fluorene-d10",
    "Phenanthrene": "Fluorene-d10",
    "Anthracene": "Fluorene-d10",
    "Fluoranthene": "Pyrene-d10",
    "Pyrene": "Pyrene-d10",
    "Benz[a]anthracene": "Pyrene-d10",
    "Chrysene": "Pyrene-d10",
    "Benzo[b]fluoranthene": "Benzo[a]pyrene-d12",
    "Benzo[k]fluoranthene": "Benzo[a]pyrene-d12",
    "Benzo[a]pyrene": "Benzo[a]pyrene-d12",
    "Indeno[1,2,3-cd]pyrene": "Benzo[a]pyrene-d12",
    "Dibenz[a,h]anthracene": "Benzo[a]pyrene-d12",
    "Benzo[g,h,i]perylene": "Benzo[a]pyrene-d12",
}

SURROGATE_SPIKE_BY_NAME = {name: spike for name, role, spike in ANALYTES if role == "Surrogate"}


IGNORE_PATTERNS_DEFAULT = [
    r"(?i)\bug/?ml\b",
    r"(?i)\bugml\b",
    r"(?i)surrog[_\s-]*mrm",
    r"(?i)\bstandard\b",
    r"(?i)\bstd\b",
    r"(?i)\bcal\b",
    r"(?i)\bblank\b",
]


def norm_name(value: Any) -> str:
    text = str(value or "").strip()
    text = re.sub(r"\s+", " ", text)
    return text.replace("–", "-").replace("—", "-").lower()


def clean_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", norm_name(value)).strip()


def strip_datafile_extension(value: Any) -> str:
    text = str(value or "").strip()
    return re.sub(r"(?i)\.d$", "", text).strip()


def strip_df_token(value: Any) -> str:
    text = strip_datafile_extension(value)
    text = re.sub(r"(?i)(?<![A-Za-z0-9])(\d+(?:\.\d+)?)\s*[x×]\b", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def df_match_keys(value: Any) -> List[str]:
    keys = []
    for candidate in [str(value or "").strip(), strip_datafile_extension(value), strip_df_token(value)]:
        if candidate:
            keys.append(candidate)
            keys.append(norm_name(candidate))
    seen = set()
    unique = []
    for key in keys:
        if key and key not in seen:
            seen.add(key)
            unique.append(key)
    return unique


def to_float_or_none(value: Any) -> Optional[float]:
    if value in (None, ""):
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def positive_or_none(value: Any) -> Optional[float]:
    val = to_float_or_none(value)
    if val is None or val <= 0:
        return None
    return val


def parse_df(sample_name: str) -> Tuple[float, str]:
    text = str(sample_name or "")
    matches = re.findall(r"(?i)(?<![A-Za-z0-9])(\d+(?:\.\d+)?)\s*[x×]\b", text)
    if matches:
        try:
            return float(matches[-1]), "Parsed from sample name"
        except ValueError:
            pass
    return 1.0, "Default 1"


def should_ignore_sample(sample_name: str, ignore_patterns: List[str] = IGNORE_PATTERNS_DEFAULT) -> bool:
    text = str(sample_name or "")
    if not text.strip():
        return True
    return any(re.search(pattern, text) for pattern in ignore_patterns)


def _find_metadata_header(ws) -> Tuple[int, Dict[str, Optional[int]]]:
    max_scan_rows = min(ws.max_row, 100)
    best_row = None
    best_cols = {}
    for row in range(1, max_scan_rows + 1):
        found = {
            "sample_name": None,
            "dilution_factor": None,
            "extract_volume_ml": None,
            "fabric_size_cm2": None,
            "fabric_mass_g": None,
            "fuel_mass_loss_g": None,
        }
        for col in range(1, ws.max_column + 1):
            compact = clean_header(ws.cell(row, col).value)
            if not compact:
                continue
            if compact in {"sample name", "sample"} or ("sample" in compact and "name" in compact):
                found["sample_name"] = col
            if compact in {"dilution factor", "df"} or ("dilution" in compact and "factor" in compact):
                found["dilution_factor"] = col
            if ("sample" in compact and ("volume" in compact or "mass" in compact)) or compact in {"extract volume ml", "sample volume ml", "volume ml"}:
                found["extract_volume_ml"] = col
            if "fabric" in compact and "size" in compact:
                found["fabric_size_cm2"] = col
            if "fabric" in compact and "mass" in compact:
                found["fabric_mass_g"] = col
            if "fuel" in compact and "mass" in compact and "loss" in compact:
                found["fuel_mass_loss_g"] = col
        if found["sample_name"] and any(found[k] for k in found if k != "sample_name"):
            best_row = row
            best_cols = found
            if found["dilution_factor"]:
                break
    if not best_row:
        raise ValueError(
            "Could not find sample metadata columns. The file must contain a sample-name column, "
            "such as 'Sample Name' or 'Sample Name (Unique)'. Optional columns include 'Dilution Factor', "
            "'Fabric size (cm^2)', 'fabric mass (g)', and 'fuel mass loss (g)'."
        )
    return best_row, best_cols


def load_sample_metadata(metadata_path: Optional[Path]) -> Dict[str, Dict[str, Any]]:
    if not metadata_path:
        return {}
    wb = load_workbook(metadata_path, data_only=True)
    ws = wb["Dilution_Factors"] if "Dilution_Factors" in wb.sheetnames else wb[wb.sheetnames[0]]
    header_row, cols = _find_metadata_header(ws)
    metadata: Dict[str, Dict[str, Any]] = {}
    for row in range(header_row + 1, ws.max_row + 1):
        sample = ws.cell(row, cols["sample_name"]).value if cols["sample_name"] else None
        if sample in (None, ""):
            continue
        rec = {
            "sample_name": str(sample).strip(),
            "dilution_factor": to_float_or_none(ws.cell(row, cols["dilution_factor"]).value) if cols["dilution_factor"] else None,
            "extract_volume_ml": to_float_or_none(ws.cell(row, cols["extract_volume_ml"]).value) if cols["extract_volume_ml"] else None,
            "fabric_size_cm2": to_float_or_none(ws.cell(row, cols["fabric_size_cm2"]).value) if cols["fabric_size_cm2"] else None,
            "fabric_mass_g": to_float_or_none(ws.cell(row, cols["fabric_mass_g"]).value) if cols["fabric_mass_g"] else None,
            "fuel_mass_loss_g": to_float_or_none(ws.cell(row, cols["fuel_mass_loss_g"]).value) if cols["fuel_mass_loss_g"] else None,
            "source": "Uploaded submission/metadata file",
        }
        for key in df_match_keys(sample):
            metadata[key] = rec
    return metadata


def lookup_sample_metadata(metadata: Optional[Dict[str, Dict[str, Any]]], *values: Any) -> Optional[Dict[str, Any]]:
    if not metadata:
        return None
    for value in values:
        for key in df_match_keys(value):
            if key in metadata:
                return metadata[key]
    return None


def find_header_rows(ws) -> Tuple[int, int]:
    for r in range(1, min(ws.max_row, 25) + 1):
        values = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 60) + 1)]
        if "Name" in values and "Final Conc." in values:
            return r - 1, r
    return 1, 2


def find_column(ws, subheader_row: int, target: str) -> Optional[int]:
    for c in range(1, ws.max_column + 1):
        if norm_name(ws.cell(subheader_row, c).value) == norm_name(target):
            return c
    return None


def find_acq_col(ws, subheader_row: int) -> Optional[int]:
    for c in range(1, ws.max_column + 1):
        label = norm_name(ws.cell(subheader_row, c).value)
        if "acq" in label and "date" in label:
            return c
    return None


def find_analyte_final_conc_cols(ws, analytes, analyte_header_row: int, subheader_row: int):
    header_analytes = {}
    for c in range(1, ws.max_column + 1):
        value = ws.cell(analyte_header_row, c).value
        if isinstance(value, str) and value.strip().lower().endswith(" results"):
            analyte = re.sub(r"\s+Results\s*$", "", value.strip(), flags=re.I)
            header_analytes[norm_name(analyte)] = c
    result = {}
    missing = []
    for analyte, _, _ in analytes:
        start_col = header_analytes.get(norm_name(analyte))
        if not start_col:
            missing.append(analyte)
            continue
        final_col = None
        for cc in range(start_col, min(ws.max_column, start_col + 10) + 1):
            if norm_name(ws.cell(subheader_row, cc).value) == "final conc.":
                final_col = cc
                break
        if final_col is None:
            missing.append(analyte)
        else:
            result[analyte] = final_col
    return result, missing


def calc_normalized_values(corrected: Any, extract_volume_ml: Any, fabric_size_cm2: Any, fabric_mass_g: Any, fuel_mass_loss_g: Any) -> Dict[str, Optional[float]]:
    corr = to_float_or_none(corrected)
    vol = positive_or_none(extract_volume_ml) or 1.0
    size = positive_or_none(fabric_size_cm2)
    mass = positive_or_none(fabric_mass_g)
    fuel = positive_or_none(fuel_mass_loss_g)
    if corr is None:
        return {"mass_ug": None, "surface_load": None, "ug_g_fabric": None, "ug_g_fuel_loss": None}
    mass_ug = corr * vol
    return {
        "mass_ug": mass_ug,
        "surface_load": mass_ug * 100 / size if size else None,
        "ug_g_fabric": mass_ug / mass if mass else None,
        "ug_g_fuel_loss": mass_ug / fuel if fuel else None,
    }


def sum_numeric(values: List[Any]) -> Optional[float]:
    numeric = [to_float_or_none(v) for v in values]
    numeric = [v for v in numeric if v is not None]
    return sum(numeric) if numeric else None


def category_metrics(s: Dict[str, Any], analytes: List[str], teq: bool = False) -> Dict[str, Optional[float]]:
    values = []
    for analyte in analytes:
        corrected = s["corrected_values"].get(analyte)
        if teq:
            factor = TEF_FACTORS.get(analyte, 0)
            corrected = corrected * factor if isinstance(corrected, (int, float)) else None
        values.append(corrected)
    corrected_sum = sum_numeric(values)
    norm = calc_normalized_values(corrected_sum, s.get("extract_volume_ml"), s.get("fabric_size_cm2"), s.get("fabric_mass_g"), s.get("fuel_mass_loss_g"))
    return {
        "corrected_conc": corrected_sum,
        "surface_load": norm["surface_load"],
        "ug_g_fabric": norm["ug_g_fabric"],
        "ug_g_fuel_loss": norm["ug_g_fuel_loss"],
    }


def extract_gcms(input_xlsx: Path, sample_metadata=None, keep_ignored=False):
    wb_in = load_workbook(input_xlsx, data_only=True, read_only=False)
    ws = wb_in[wb_in.sheetnames[0]]
    analyte_header_row, subheader_row = find_header_rows(ws)
    name_col = find_column(ws, subheader_row, "Name") or 3
    datafile_col = find_column(ws, subheader_row, "Data File")
    acq_col = find_acq_col(ws, subheader_row)
    analyte_cols, missing = find_analyte_final_conc_cols(ws, ANALYTES, analyte_header_row, subheader_row)
    samples = []
    for row_idx in range(subheader_row + 1, ws.max_row + 1):
        name = ws.cell(row_idx, name_col).value
        if name is None or str(name).strip() == "":
            continue
        ignored = should_ignore_sample(str(name))
        if ignored and not keep_ignored:
            continue
        data_file = ws.cell(row_idx, datafile_col).value if datafile_col else None
        df, df_source = parse_df(str(name))
        meta_rec = lookup_sample_metadata(sample_metadata, name, data_file)
        extract_volume_ml = 1.0
        fabric_size_cm2 = None
        fabric_mass_g = None
        fuel_mass_loss_g = None
        parameter_source = "Default/blank"
        if meta_rec:
            if meta_rec.get("dilution_factor") is not None:
                df = meta_rec.get("dilution_factor")
                df_source = "Manual/submission form override"
            extract_volume_ml = meta_rec.get("extract_volume_ml") or 1.0
            fabric_size_cm2 = meta_rec.get("fabric_size_cm2")
            fabric_mass_g = meta_rec.get("fabric_mass_g")
            fuel_mass_loss_g = meta_rec.get("fuel_mass_loss_g")
            parameter_source = meta_rec.get("source", "Uploaded submission/metadata file")
        record = {
            "row": row_idx,
            "sample_name": str(name),
            "data_file": data_file,
            "acq_datetime": ws.cell(row_idx, acq_col).value if acq_col else None,
            "dilution_factor": float(df) if df not in (None, "") else 1.0,
            "df_source": df_source,
            "extract_volume_ml": float(extract_volume_ml) if extract_volume_ml not in (None, "") else 1.0,
            "fabric_size_cm2": fabric_size_cm2,
            "fabric_mass_g": fabric_mass_g,
            "fuel_mass_loss_g": fuel_mass_loss_g,
            "parameter_source": parameter_source,
            "ignored": ignored,
            "values": {},
            "corrected_values": {},
            "normalized_values": {},
            "summary_metrics": {},
            "teq_metrics": {},
        }
        for analyte, _, _ in ANALYTES:
            col = analyte_cols.get(analyte)
            value = ws.cell(row_idx, col).value if col else None
            if isinstance(value, str):
                try:
                    value = float(value)
                except ValueError:
                    pass
            record["values"][analyte] = value
            corrected = value * record["dilution_factor"] if isinstance(value, (int, float)) else None
            record["corrected_values"][analyte] = corrected
            if analyte in TARGET_16_PAHS:
                record["normalized_values"][analyte] = calc_normalized_values(corrected, record["extract_volume_ml"], fabric_size_cm2, fabric_mass_g, fuel_mass_loss_g)
        for label, group in [("Total PAHs", TARGET_16_PAHS), ("LMW PAHs", LMW_PAHS), ("HMW PAHs", HMW_PAHS)]:
            record["summary_metrics"][label] = category_metrics(record, group, teq=False)
            record["teq_metrics"][label] = category_metrics(record, group, teq=True)
        samples.append(record)
    meta = {
        "input_file": str(input_xlsx),
        "sheet_name": ws.title,
        "analyte_header_row": analyte_header_row,
        "subheader_row": subheader_row,
        "name_col": get_column_letter(name_col),
        "data_file_col": get_column_letter(datafile_col) if datafile_col else None,
        "acq_col": get_column_letter(acq_col) if acq_col else None,
        "analyte_final_conc_cols": {k: get_column_letter(v) for k, v in analyte_cols.items()},
        "missing_analytes": missing,
        "extracted_samples": len(samples),
        "ignored_patterns": IGNORE_PATTERNS_DEFAULT,
        "normalization_formula": "mass_ug = corrected_conc * extract_volume_ml; surface_load = mass_ug * 100 / fabric_size_cm2; ug_g_fabric = mass_ug / fabric_mass_g; ug_g_fuel_loss = mass_ug / fuel_mass_loss_g",
    }
    return samples, meta



def surrogate_recovery_pct_isu(sample: Dict[str, Any], surrogate: str) -> Optional[float]:
    """Return surrogate recovery percentage for an ISU-format sample."""
    spike = SURROGATE_SPIKE_BY_NAME.get(surrogate)
    corrected = sample.get("corrected_values", {}).get(surrogate)
    if isinstance(corrected, (int, float)) and spike:
        return corrected / spike * 100.0
    return None


def recovery_corrected_conc_isu(sample: Dict[str, Any], analyte: str) -> Optional[float]:
    """DF-corrected concentration adjusted to 100% surrogate recovery."""
    corrected = sample.get("corrected_values", {}).get(analyte)
    surrogate = RECOVERY_SURROGATE_MAP_ISU.get(analyte)
    recovery = surrogate_recovery_pct_isu(sample, surrogate) if surrogate else None
    if isinstance(corrected, (int, float)) and isinstance(recovery, (int, float)) and recovery > 0:
        return corrected * 100.0 / recovery
    return None


def recovery_correction_factor_isu(sample: Dict[str, Any], analyte: str) -> Optional[float]:
    surrogate = RECOVERY_SURROGATE_MAP_ISU.get(analyte)
    recovery = surrogate_recovery_pct_isu(sample, surrogate) if surrogate else None
    if isinstance(recovery, (int, float)) and recovery > 0:
        return 100.0 / recovery
    return None


def category_metrics_recovery_isu(s: Dict[str, Any], analytes: List[str], teq: bool = False) -> Dict[str, Optional[float]]:
    values = []
    for analyte in analytes:
        val = recovery_corrected_conc_isu(s, analyte)
        if teq and isinstance(val, (int, float)):
            val = val * TEF_FACTORS.get(analyte, 0)
        values.append(val)
    corrected_sum = sum_numeric(values)
    norm = calc_normalized_values(
        corrected_sum,
        s.get("extract_volume_ml"),
        s.get("fabric_size_cm2"),
        s.get("fabric_mass_g"),
        s.get("fuel_mass_loss_g"),
    )
    return {
        "corrected_conc": corrected_sum,
        "surface_load": norm["surface_load"],
        "ug_g_fabric": norm["ug_g_fabric"],
        "ug_g_fuel_loss": norm["ug_g_fuel_loss"],
    }


def _style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True, color="1F2937")
    thin = Side(style="thin", color="D9E2EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
    for col in range(1, ws.max_column + 1):
        max_len = 8
        for row in range(1, min(ws.max_row, 120) + 1):
            value = ws.cell(row, col).value
            max_len = max(max_len, min(len(str(value)) if value is not None else 0, 55))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 48)
    ws.freeze_panes = "A2"


def write_processed_workbook(samples, meta, output_path: Path) -> None:
    wb = Workbook()
    ws_info = wb.active
    ws_info.title = "Instructions"
    ws_samples = wb.create_sheet("Sample_Info")
    ws_wide = wb.create_sheet("Wide_Data")
    ws_summary = wb.create_sheet("Summary_Metrics")
    ws_long = wb.create_sheet("Long_Data")
    ws_qc = wb.create_sheet("Recovery_QC")
    ws_config = wb.create_sheet("Config_Analytes")
    ws_run = wb.create_sheet("Run_Info")

    ws_info.append(["GC/MS PAH Extraction Output v5 - ISU format"])
    ws_info.append(["Corrected concentration = Raw Final Conc. × Dilution Factor"])
    ws_info.append(["Surrogate recovery % = Surrogate Corrected Conc. ÷ Surrogate Spike Conc. × 100"])
    ws_info.append(["Recovery-Corrected Conc. = Corrected Conc. × 100 ÷ Assigned Surrogate Recovery %"])
    ws_info.append(["Recovery-corrected columns are added in addition to, not instead of, the original DF-corrected columns."])
    ws_info.append(["Analyte mass (ug) = concentration (ug/mL) × Extract Volume (mL). Default Extract Volume is 1 mL if not provided."])
    ws_info.append(["Surface load (ug/100 cm^2) = Analyte mass × 100 / Fabric size (cm^2)"])
    ws_info.append(["ug/g fabric = Analyte mass / Fabric mass (g)"])
    ws_info.append(["ug/g fuel mass loss = Analyte mass / Fuel mass loss (g)"])
    ws_info.append(["TEQ values are calculated using the TEF factors listed in Config_Analytes."])

    ws_samples.append(["Sample Name", "Data File", "Acquisition Date/Time", "Dilution Factor", "DF Source", "Extract Volume (mL)", "Fabric Size (cm^2)", "Fabric Mass (g)", "Fuel Mass Loss (g)", "Parameter Source", "Original GC/MS Row"])
    for s in samples:
        ws_samples.append([s["sample_name"], s["data_file"], s["acq_datetime"], s["dilution_factor"], s["df_source"], s["extract_volume_ml"], s["fabric_size_cm2"], s["fabric_mass_g"], s["fuel_mass_loss_g"], s["parameter_source"], s["row"]])

    wide_headers = ["Sample Name", "Data File", "Acquisition Date/Time", "Dilution Factor", "DF Source", "Extract Volume (mL)", "Fabric Size (cm^2)", "Fabric Mass (g)", "Fuel Mass Loss (g)"]
    for analyte, role, spike in ANALYTES:
        wide_headers.extend([f"{analyte} Raw Final Conc.", f"{analyte} Corrected Conc. (ug/mL)"])
        if analyte in TARGET_16_PAHS:
            wide_headers.extend([
                f"{analyte} surface load (ug/100 cm^2)",
                f"{analyte} ug/g fabric",
                f"{analyte} ug/g mass loss of fuel",
                f"{analyte} Assigned Recovery Surrogate",
                f"{analyte} Surrogate Recovery %",
                f"{analyte} Recovery Correction Factor",
                f"{analyte} Recovery-Corrected Conc. (ug/mL)",
                f"{analyte} Recovery-Corrected surface load (ug/100 cm^2)",
                f"{analyte} Recovery-Corrected ug/g fabric",
                f"{analyte} Recovery-Corrected ug/g mass loss of fuel",
            ])
        if role == "Surrogate":
            wide_headers.append(f"{analyte} Recovery %")
    for label in ["Total PAHs", "LMW PAHs", "HMW PAHs"]:
        wide_headers.extend([f"{label} Corrected Conc. (ug/mL)", f"{label} surface load (ug/100 cm^2)", f"{label} ug/g fabric", f"{label} ug/g mass loss of fuel"])
    for label in ["Total PAHs", "LMW PAHs", "HMW PAHs"]:
        wide_headers.extend([f"{label} TEQ Corrected Conc. (ug/mL)", f"{label} TEQ surface load (ug/100 cm^2)", f"{label} TEQ ug/g fabric", f"{label} TEQ ug/g mass loss of fuel"])
    for label in ["Total PAHs", "LMW PAHs", "HMW PAHs"]:
        wide_headers.extend([f"{label} Recovery-Corrected Conc. (ug/mL)", f"{label} Recovery-Corrected surface load (ug/100 cm^2)", f"{label} Recovery-Corrected ug/g fabric", f"{label} Recovery-Corrected ug/g mass loss of fuel"])
    for label in ["Total PAHs", "LMW PAHs", "HMW PAHs"]:
        wide_headers.extend([f"{label} Recovery-Corrected TEQ Conc. (ug/mL)", f"{label} Recovery-Corrected TEQ surface load (ug/100 cm^2)", f"{label} Recovery-Corrected TEQ ug/g fabric", f"{label} Recovery-Corrected TEQ ug/g mass loss of fuel"])
    ws_wide.append(wide_headers)

    for s in samples:
        row = [s["sample_name"], s["data_file"], s["acq_datetime"], s["dilution_factor"], s["df_source"], s["extract_volume_ml"], s["fabric_size_cm2"], s["fabric_mass_g"], s["fuel_mass_loss_g"]]
        for analyte, role, spike in ANALYTES:
            raw = s["values"].get(analyte)
            corrected = s["corrected_values"].get(analyte)
            row.extend([raw, corrected])
            if analyte in TARGET_16_PAHS:
                norm = s["normalized_values"].get(analyte, {})
                assigned = RECOVERY_SURROGATE_MAP_ISU.get(analyte)
                rec_pct = surrogate_recovery_pct_isu(s, assigned) if assigned else None
                rec_factor = recovery_correction_factor_isu(s, analyte)
                rec_corr = recovery_corrected_conc_isu(s, analyte)
                rec_norm = calc_normalized_values(rec_corr, s.get("extract_volume_ml"), s.get("fabric_size_cm2"), s.get("fabric_mass_g"), s.get("fuel_mass_loss_g"))
                row.extend([
                    norm.get("surface_load"), norm.get("ug_g_fabric"), norm.get("ug_g_fuel_loss"),
                    assigned, rec_pct, rec_factor, rec_corr,
                    rec_norm.get("surface_load"), rec_norm.get("ug_g_fabric"), rec_norm.get("ug_g_fuel_loss"),
                ])
            if role == "Surrogate":
                recovery = corrected / spike * 100 if isinstance(corrected, (int, float)) and spike else None
                row.append(recovery)
        for label in ["Total PAHs", "LMW PAHs", "HMW PAHs"]:
            m = s["summary_metrics"][label]
            row.extend([m["corrected_conc"], m["surface_load"], m["ug_g_fabric"], m["ug_g_fuel_loss"]])
        for label in ["Total PAHs", "LMW PAHs", "HMW PAHs"]:
            m = s["teq_metrics"][label]
            row.extend([m["corrected_conc"], m["surface_load"], m["ug_g_fabric"], m["ug_g_fuel_loss"]])
        for label, group in [("Total PAHs", TARGET_16_PAHS), ("LMW PAHs", LMW_PAHS), ("HMW PAHs", HMW_PAHS)]:
            m = category_metrics_recovery_isu(s, group, teq=False)
            row.extend([m["corrected_conc"], m["surface_load"], m["ug_g_fabric"], m["ug_g_fuel_loss"]])
        for label, group in [("Total PAHs", TARGET_16_PAHS), ("LMW PAHs", LMW_PAHS), ("HMW PAHs", HMW_PAHS)]:
            m = category_metrics_recovery_isu(s, group, teq=True)
            row.extend([m["corrected_conc"], m["surface_load"], m["ug_g_fabric"], m["ug_g_fuel_loss"]])
        ws_wide.append(row)

    ws_summary.append([
        "Sample Name", "Category",
        "Corrected Conc. (ug/mL)", "surface load (ug/100 cm^2)", "ug/g fabric", "ug/g mass loss of fuel",
        "TEQ Corrected Conc. (ug/mL)", "TEQ surface load (ug/100 cm^2)", "TEQ ug/g fabric", "TEQ ug/g mass loss of fuel",
        "Recovery-Corrected Conc. (ug/mL)", "Recovery-Corrected surface load (ug/100 cm^2)", "Recovery-Corrected ug/g fabric", "Recovery-Corrected ug/g mass loss of fuel",
        "Recovery-Corrected TEQ Conc. (ug/mL)", "Recovery-Corrected TEQ surface load (ug/100 cm^2)", "Recovery-Corrected TEQ ug/g fabric", "Recovery-Corrected TEQ ug/g mass loss of fuel",
    ])
    for s in samples:
        for label, group in [("Total PAHs", TARGET_16_PAHS), ("LMW PAHs", LMW_PAHS), ("HMW PAHs", HMW_PAHS)]:
            m = s["summary_metrics"][label]
            t = s["teq_metrics"][label]
            r = category_metrics_recovery_isu(s, group, teq=False)
            rt = category_metrics_recovery_isu(s, group, teq=True)
            ws_summary.append([s["sample_name"], label, m["corrected_conc"], m["surface_load"], m["ug_g_fabric"], m["ug_g_fuel_loss"], t["corrected_conc"], t["surface_load"], t["ug_g_fabric"], t["ug_g_fuel_loss"], r["corrected_conc"], r["surface_load"], r["ug_g_fabric"], r["ug_g_fuel_loss"], rt["corrected_conc"], rt["surface_load"], rt["ug_g_fabric"], rt["ug_g_fuel_loss"]])

    ws_long.append([
        "Sample Name", "Data File", "Acquisition Date/Time", "Dilution Factor", "Extract Volume (mL)", "Fabric Size (cm^2)", "Fabric Mass (g)", "Fuel Mass Loss (g)",
        "Analyte", "Role", "Raw Final Conc.", "Corrected Conc. (ug/mL)", "Analyte mass (ug)", "surface load (ug/100 cm^2)", "ug/g fabric", "ug/g mass loss of fuel",
        "Assigned Recovery Surrogate", "Surrogate Recovery %", "Recovery Correction Factor", "Recovery-Corrected Conc. (ug/mL)", "Recovery-Corrected mass (ug)", "Recovery-Corrected surface load (ug/100 cm^2)", "Recovery-Corrected ug/g fabric", "Recovery-Corrected ug/g mass loss of fuel",
        "TEF", "TEQ Corrected Conc. (ug/mL)", "TEQ surface load (ug/100 cm^2)", "TEQ ug/g fabric", "TEQ ug/g mass loss of fuel", "TEQ Recovery-Corrected Conc. (ug/mL)", "TEQ Recovery-Corrected surface load (ug/100 cm^2)", "TEQ Recovery-Corrected ug/g fabric", "TEQ Recovery-Corrected ug/g mass loss of fuel", "Surrogate Spike", "Recovery %",
    ])
    for s in samples:
        for analyte, role, spike in ANALYTES:
            raw = s["values"].get(analyte)
            corrected = s["corrected_values"].get(analyte)
            norm = calc_normalized_values(corrected, s.get("extract_volume_ml"), s.get("fabric_size_cm2"), s.get("fabric_mass_g"), s.get("fuel_mass_loss_g")) if analyte in TARGET_16_PAHS else {"mass_ug": None, "surface_load": None, "ug_g_fabric": None, "ug_g_fuel_loss": None}
            assigned = RECOVERY_SURROGATE_MAP_ISU.get(analyte) if analyte in TARGET_16_PAHS else None
            rec_pct = surrogate_recovery_pct_isu(s, assigned) if assigned else None
            rec_factor = recovery_correction_factor_isu(s, analyte) if analyte in TARGET_16_PAHS else None
            rec_corr = recovery_corrected_conc_isu(s, analyte) if analyte in TARGET_16_PAHS else None
            rec_norm = calc_normalized_values(rec_corr, s.get("extract_volume_ml"), s.get("fabric_size_cm2"), s.get("fabric_mass_g"), s.get("fuel_mass_loss_g")) if analyte in TARGET_16_PAHS else {"mass_ug": None, "surface_load": None, "ug_g_fabric": None, "ug_g_fuel_loss": None}
            tef = TEF_FACTORS.get(analyte)
            teq_corr = corrected * tef if isinstance(corrected, (int, float)) and tef is not None else None
            teq_norm = calc_normalized_values(teq_corr, s.get("extract_volume_ml"), s.get("fabric_size_cm2"), s.get("fabric_mass_g"), s.get("fuel_mass_loss_g")) if teq_corr is not None else {"surface_load": None, "ug_g_fabric": None, "ug_g_fuel_loss": None}
            teq_rec_corr = rec_corr * tef if isinstance(rec_corr, (int, float)) and tef is not None else None
            teq_rec_norm = calc_normalized_values(teq_rec_corr, s.get("extract_volume_ml"), s.get("fabric_size_cm2"), s.get("fabric_mass_g"), s.get("fuel_mass_loss_g")) if teq_rec_corr is not None else {"surface_load": None, "ug_g_fabric": None, "ug_g_fuel_loss": None}
            recovery = corrected / spike * 100 if role == "Surrogate" and isinstance(corrected, (int, float)) and spike else None
            ws_long.append([s["sample_name"], s["data_file"], s["acq_datetime"], s["dilution_factor"], s["extract_volume_ml"], s["fabric_size_cm2"], s["fabric_mass_g"], s["fuel_mass_loss_g"], analyte, role, raw, corrected, norm.get("mass_ug"), norm.get("surface_load"), norm.get("ug_g_fabric"), norm.get("ug_g_fuel_loss"), assigned, rec_pct, rec_factor, rec_corr, rec_norm.get("mass_ug"), rec_norm.get("surface_load"), rec_norm.get("ug_g_fabric"), rec_norm.get("ug_g_fuel_loss"), tef, teq_corr, teq_norm.get("surface_load"), teq_norm.get("ug_g_fabric"), teq_norm.get("ug_g_fuel_loss"), teq_rec_corr, teq_rec_norm.get("surface_load"), teq_rec_norm.get("ug_g_fabric"), teq_rec_norm.get("ug_g_fuel_loss"), spike, recovery])

    ws_qc.append(["Sample Name", "Surrogate", "Corrected Conc. (ug/mL)", "Spike Conc. (ug/mL)", "Recovery %", "QC Flag"])
    for s in samples:
        for analyte, role, spike in ANALYTES:
            if role != "Surrogate":
                continue
            corrected = s["corrected_values"].get(analyte)
            recovery = corrected / spike * 100 if isinstance(corrected, (int, float)) and spike else None
            flag = "OK" if isinstance(recovery, (int, float)) and 50 <= recovery <= 150 else "Check"
            ws_qc.append([s["sample_name"], analyte, corrected, spike, recovery, flag])

    ws_config.append(["Analyte", "Role", "Group", "Surrogate Spike Conc. (ug/mL)", "TEF", "Assigned Recovery Surrogate"])
    for analyte, role, spike in ANALYTES:
        group = "LMW PAH" if analyte in LMW_PAHS else "HMW PAH" if analyte in HMW_PAHS else "Surrogate"
        ws_config.append([analyte, role, group, spike, TEF_FACTORS.get(analyte), RECOVERY_SURROGATE_MAP_ISU.get(analyte)])

    ws_run.append(["Field", "Value"])
    for key, value in meta.items():
        ws_run.append([key, str(value)])
    ws_run.append(["Recovery correction formula", "Recovery-Corrected Conc. = Corrected Conc. × 100 / Assigned Surrogate Recovery %"])
    ws_run.append(["Created", datetime.datetime.now().isoformat(timespec="seconds")])

    for ws in wb.worksheets:
        _style_sheet(ws)
    for ws in [ws_wide, ws_summary, ws_long, ws_qc]:
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = "0.0000"
    if ws_qc.max_row > 1:
        red_fill = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")
        ws_qc.conditional_formatting.add(f"F2:F{ws_qc.max_row}", CellIsRule(operator="equal", formula=['"Check"'], fill=red_fill))
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="Extract GC/MS PAH Final Conc. values, apply dilution factors, and compute PAH normalization metrics.")
    parser.add_argument("input_xlsx", type=Path)
    parser.add_argument("--df-template", type=Path, default=None)
    parser.add_argument("--metadata", type=Path, default=None)
    parser.add_argument("--output", type=Path, default=None)
    args = parser.parse_args()
    metadata_path = args.metadata or args.df_template
    sample_metadata = load_sample_metadata(metadata_path) if metadata_path else {}
    samples, meta = extract_gcms(args.input_xlsx, sample_metadata=sample_metadata)
    output = args.output or args.input_xlsx.with_name(args.input_xlsx.stem + "_processed.xlsx")
    write_processed_workbook(samples, meta, output)
    print(f"Processed {len(samples)} samples.")
    print(f"Sample metadata entries loaded: {len(sample_metadata)}")
    print(f"Output written to: {output}")


if __name__ == "__main__":
    main()
