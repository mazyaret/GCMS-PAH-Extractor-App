#!/usr/bin/env python3
"""
GC/MS PAH Extractor - NoTableFix v2.1

Updates in v2:
- Keeps Excel table objects disabled to avoid Excel repair warnings.
- Accepts either the simple DF template OR the lab sample submission form.
- Matches DF by exact sample name, GC/MS data-file name, or base name after removing labels like 8X.
- For the lab submission form, it automatically finds columns like:
  "Sample Name (Unique)" and "Dilution factor", even when the header row is not row 1.
"""

from __future__ import annotations

import argparse
import datetime
import re
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.comments import Comment

ANALYTES: List[Tuple[str, str, Optional[float]]] = [
    ("Nitrobenzene-d5", "Surrogate", 0.5),
    ("Naphthalene", "Target", None),
    ("2-fluorobiphenyl", "Surrogate", 0.5),
    ("Acenaphthylene", "Target", None),
    ("Acenaphthene", "Target", None),
    ("Fluorene-d10", "Surrogate", 0.5),
    ("Fluorene", "Target", None),
    ("2,4,6 Tribromophenol", "Surrogate", 0.5),
    ("Phenanthrene", "Target", None),
    ("Anthracene", "Target", None),
    ("Fluoranthene", "Target", None),
    ("Pyrene-d10", "Surrogate", 0.5),
    ("Pyrene", "Target", None),
    ("P-terphenyl-d14", "Surrogate", 0.5),
    ("Benz[a]anthracene", "Target", None),
    ("Chrysene", "Target", None),
    ("Benzo[b]fluoranthene", "Target", None),
    ("Benzo[k]fluoranthene", "Target", None),
    ("Benzo[a]pyrene-d12", "Surrogate", 0.5),
    ("Benzo[a]pyrene", "Target", None),
    ("Indeno[1,2,3-cd]pyrene", "Target", None),
    ("Dibenz[a,h]anthracene", "Target", None),
    ("Benzo[g,h,i]perylene", "Target", None),
]

IGNORE_PATTERNS_DEFAULT = [
    r"(?i)\bug/?ml\b",
    r"(?i)\bugml\b",
    r"(?i)surrog[_\s-]*mrm",
    r"(?i)\bstandard\b",
    r"(?i)\bstd\b",
    r"(?i)\bcal\b",
    r"(?i)\bblank\b",
]


def norm_name(value: Any) -> str:
    text = str(value or "").strip()
    text = re.sub(r"\s+", " ", text)
    return text.replace("–", "-").replace("—", "-").lower()


def strip_datafile_extension(value: Any) -> str:
    text = str(value or "").strip()
    return re.sub(r"(?i)\.d$", "", text).strip()


def strip_df_token(value: Any) -> str:
    text = strip_datafile_extension(value)
    text = re.sub(r"(?i)(?<![A-Za-z0-9])(\d+(?:\.\d+)?)\s*[x×]\b", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def df_match_keys(value: Any) -> List[str]:
    keys = []
    for candidate in [str(value or "").strip(), strip_datafile_extension(value), strip_df_token(value)]:
        if candidate:
            keys.append(candidate)
            keys.append(norm_name(candidate))
    seen = set()
    unique = []
    for key in keys:
        if key and key not in seen:
            seen.add(key)
            unique.append(key)
    return unique


def lookup_df_override(df_overrides: Optional[Dict[str, float]], *values: Any) -> Optional[float]:
    if not df_overrides:
        return None
    for value in values:
        for key in df_match_keys(value):
            if key in df_overrides:
                return df_overrides[key]
    return None


def parse_df(sample_name: str) -> Tuple[float, str]:
    text = str(sample_name or "")
    matches = re.findall(r"(?i)(?<![A-Za-z0-9])(\d+(?:\.\d+)?)\s*[x×]\b", text)
    if matches:
        try:
            return float(matches[-1]), "Parsed from sample name"
        except ValueError:
            pass
    return 1.0, "Default 1"


def should_ignore_sample(sample_name: str, ignore_patterns: List[str] = IGNORE_PATTERNS_DEFAULT) -> bool:
    text = str(sample_name or "")
    if not text.strip():
        return True
    return any(re.search(pattern, text) for pattern in ignore_patterns)


def _find_df_header(ws) -> Tuple[int, int, int]:
    max_scan_rows = min(ws.max_row, 100)
    for row in range(1, max_scan_rows + 1):
        sample_col = None
        df_col = None
        for col in range(1, ws.max_column + 1):
            label = norm_name(ws.cell(row, col).value)
            if not label:
                continue
            compact = re.sub(r"[^a-z0-9]+", " ", label).strip()
            if label in {"sample name", "sample"} or ("sample" in compact and "name" in compact):
                sample_col = col
            if label in {"dilution factor", "df"} or ("dilution" in compact and "factor" in compact):
                df_col = col
        if sample_col and df_col:
            return row, sample_col, df_col
    raise ValueError(
        "Could not find DF columns. The DF file must contain a sample-name column "
        "such as 'Sample Name' or 'Sample Name (Unique)' and a DF column such as 'Dilution Factor'."
    )


def load_df_overrides(df_template_path: Optional[Path]) -> Dict[str, float]:
    if not df_template_path:
        return {}
    wb = load_workbook(df_template_path, data_only=True)
    if "Dilution_Factors" in wb.sheetnames:
        ws = wb["Dilution_Factors"]
    else:
        ws = wb[wb.sheetnames[0]]

    header_row, sample_col, df_col = _find_df_header(ws)

    overrides: Dict[str, float] = {}
    for row in range(header_row + 1, ws.max_row + 1):
        sample = ws.cell(row, sample_col).value
        df = ws.cell(row, df_col).value
        if sample in (None, "") or df in (None, ""):
            continue
        try:
            df_float = float(df)
        except (TypeError, ValueError):
            continue
        for key in df_match_keys(sample):
            overrides[key] = df_float
    return overrides


def find_header_rows(ws) -> Tuple[int, int]:
    for r in range(1, min(ws.max_row, 25) + 1):
        values = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 60) + 1)]
        if "Name" in values and "Final Conc." in values:
            return r - 1, r
    return 1, 2


def find_column(ws, subheader_row: int, target: str) -> Optional[int]:
    for c in range(1, ws.max_column + 1):
        if norm_name(ws.cell(subheader_row, c).value) == norm_name(target):
            return c
    return None


def find_acq_col(ws, subheader_row: int) -> Optional[int]:
    for c in range(1, ws.max_column + 1):
        label = norm_name(ws.cell(subheader_row, c).value)
        if "acq" in label and "date" in label:
            return c
    return None


def find_analyte_final_conc_cols(ws, analytes, analyte_header_row: int, subheader_row: int):
    header_analytes = {}
    for c in range(1, ws.max_column + 1):
        value = ws.cell(analyte_header_row, c).value
        if isinstance(value, str) and value.strip().lower().endswith(" results"):
            analyte = re.sub(r"\s+Results\s*$", "", value.strip(), flags=re.I)
            header_analytes[norm_name(analyte)] = c

    result = {}
    missing = []
    for analyte, _, _ in analytes:
        start_col = header_analytes.get(norm_name(analyte))
        if not start_col:
            missing.append(analyte)
            continue
        final_col = None
        for cc in range(start_col, min(ws.max_column, start_col + 10) + 1):
            if norm_name(ws.cell(subheader_row, cc).value) == "final conc.":
                final_col = cc
                break
        if final_col is None:
            missing.append(analyte)
        else:
            result[analyte] = final_col
    return result, missing


def extract_gcms(input_xlsx: Path, df_overrides=None, keep_ignored=False):
    wb_in = load_workbook(input_xlsx, data_only=True, read_only=False)
    ws = wb_in[wb_in.sheetnames[0]]

    analyte_header_row, subheader_row = find_header_rows(ws)
    name_col = find_column(ws, subheader_row, "Name") or 3
    datafile_col = find_column(ws, subheader_row, "Data File")
    acq_col = find_acq_col(ws, subheader_row)
    analyte_cols, missing = find_analyte_final_conc_cols(ws, ANALYTES, analyte_header_row, subheader_row)

    samples = []
    for row_idx in range(subheader_row + 1, ws.max_row + 1):
        name = ws.cell(row_idx, name_col).value
        if name is None or str(name).strip() == "":
            continue

        ignored = should_ignore_sample(str(name))
        if ignored and not keep_ignored:
            continue

        data_file = ws.cell(row_idx, datafile_col).value if datafile_col else None
        df, df_source = parse_df(str(name))
        override_df = lookup_df_override(df_overrides, name, data_file)
        if override_df is not None:
            df = override_df
            df_source = "Manual/submission form override"

        record = {
            "row": row_idx,
            "sample_name": str(name),
            "data_file": data_file,
            "acq_datetime": ws.cell(row_idx, acq_col).value if acq_col else None,
            "dilution_factor": float(df) if df not in (None, "") else 1.0,
            "df_source": df_source,
            "ignored": ignored,
            "values": {},
        }

        for analyte, _, _ in ANALYTES:
            col = analyte_cols.get(analyte)
            value = ws.cell(row_idx, col).value if col else None
            if isinstance(value, str):
                try:
                    value = float(value)
                except ValueError:
                    pass
            record["values"][analyte] = value
        samples.append(record)

    meta = {
        "input_file": str(input_xlsx),
        "sheet_name": ws.title,
        "analyte_header_row": analyte_header_row,
        "subheader_row": subheader_row,
        "name_col": get_column_letter(name_col),
        "data_file_col": get_column_letter(datafile_col) if datafile_col else None,
        "acq_col": get_column_letter(acq_col) if acq_col else None,
        "analyte_final_conc_cols": {k: get_column_letter(v) for k, v in analyte_cols.items()},
        "missing_analytes": missing,
        "extracted_samples": len(samples),
        "ignored_patterns": IGNORE_PATTERNS_DEFAULT,
    }
    return samples, meta


def add_table(ws, table_name: str, start_row: int, start_col: int, end_row: int, end_col: int) -> None:
    return


def _style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True, color="1F2937")
    thin = Side(style="thin", color="D9E2EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
    for col in range(1, ws.max_column + 1):
        max_len = 8
        for row in range(1, min(ws.max_row, 120) + 1):
            value = ws.cell(row, col).value
            max_len = max(max_len, min(len(str(value)) if value is not None else 0, 45))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 45)
    ws.freeze_panes = "A2"


def write_processed_workbook(samples, meta, output_path: Path) -> None:
    wb = Workbook()
    ws_info = wb.active
    ws_info.title = "Instructions"
    ws_samples = wb.create_sheet("Sample_Info")
    ws_wide = wb.create_sheet("Wide_Data")
    ws_long = wb.create_sheet("Long_Data")
    ws_qc = wb.create_sheet("Recovery_QC")
    ws_config = wb.create_sheet("Config_Analytes")
    ws_run = wb.create_sheet("Run_Info")

    ws_info.append(["GC/MS PAH Extraction Output"])
    ws_info.append(["Corrected concentration = Raw Final Conc. × Dilution Factor"])
    ws_info.append(["Surrogate recovery % = Corrected surrogate concentration ÷ surrogate spike concentration"])
    ws_info.append(["Use Wide_Data for manual review and Long_Data for R/Python/statistics."])

    ws_samples.append(["Sample Name", "Data File", "Acquisition Date/Time", "Dilution Factor", "DF Source", "Original GC/MS Row"])
    for s in samples:
        ws_samples.append([s["sample_name"], s["data_file"], s["acq_datetime"], s["dilution_factor"], s["df_source"], s["row"]])

    wide_headers = ["Sample Name", "Data File", "Acquisition Date/Time", "Dilution Factor", "DF Source"]
    for analyte, role, spike in ANALYTES:
        wide_headers.extend([f"{analyte} Raw Final Conc.", f"{analyte} Corrected Conc."])
        if role == "Surrogate":
            wide_headers.append(f"{analyte} Recovery %")
    ws_wide.append(wide_headers)

    for s in samples:
        row = [s["sample_name"], s["data_file"], s["acq_datetime"], s["dilution_factor"], s["df_source"]]
        for analyte, role, spike in ANALYTES:
            raw = s["values"].get(analyte)
            corrected = raw * s["dilution_factor"] if isinstance(raw, (int, float)) else None
            row.extend([raw, corrected])
            if role == "Surrogate":
                recovery = corrected / spike * 100 if isinstance(corrected, (int, float)) and spike else None
                row.append(recovery)
        ws_wide.append(row)

    ws_long.append(["Sample Name", "Data File", "Acquisition Date/Time", "Dilution Factor", "DF Source", "Analyte", "Role", "Raw Final Conc.", "Corrected Conc.", "Surrogate Spike", "Recovery %"])
    for s in samples:
        for analyte, role, spike in ANALYTES:
            raw = s["values"].get(analyte)
            corrected = raw * s["dilution_factor"] if isinstance(raw, (int, float)) else None
            recovery = corrected / spike * 100 if role == "Surrogate" and isinstance(corrected, (int, float)) and spike else None
            ws_long.append([s["sample_name"], s["data_file"], s["acq_datetime"], s["dilution_factor"], s["df_source"], analyte, role, raw, corrected, spike, recovery])

    ws_qc.append(["Sample Name", "Surrogate", "Corrected Conc.", "Spike Conc.", "Recovery %", "QC Flag"])
    for s in samples:
        for analyte, role, spike in ANALYTES:
            if role != "Surrogate":
                continue
            raw = s["values"].get(analyte)
            corrected = raw * s["dilution_factor"] if isinstance(raw, (int, float)) else None
            recovery = corrected / spike * 100 if isinstance(corrected, (int, float)) and spike else None
            flag = "OK" if isinstance(recovery, (int, float)) and 50 <= recovery <= 150 else "Check"
            ws_qc.append([s["sample_name"], analyte, corrected, spike, recovery, flag])

    ws_config.append(["Analyte", "Role", "Surrogate Spike Conc. (ug/mL)"])
    for analyte, role, spike in ANALYTES:
        ws_config.append([analyte, role, spike])

    ws_run.append(["Field", "Value"])
    for key, value in meta.items():
        ws_run.append([key, str(value)])
    ws_run.append(["Created", datetime.datetime.now().isoformat(timespec="seconds")])

    for ws in wb.worksheets:
        _style_sheet(ws)

    for ws in [ws_wide, ws_long, ws_qc]:
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = "0.0000"

    if ws_qc.max_row > 1:
        red_fill = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")
        ws_qc.conditional_formatting.add(f"F2:F{ws_qc.max_row}", CellIsRule(operator="equal", formula=['"Check"'], fill=red_fill))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="Extract GC/MS PAH Final Conc. values and apply dilution factors.")
    parser.add_argument("input_xlsx", type=Path)
    parser.add_argument("--df-template", type=Path, default=None)
    parser.add_argument("--output", type=Path, default=None)
    args = parser.parse_args()

    df_overrides = load_df_overrides(args.df_template)
    samples, meta = extract_gcms(args.input_xlsx, df_overrides=df_overrides)
    output = args.output or args.input_xlsx.with_name(args.input_xlsx.stem + "_processed.xlsx")
    write_processed_workbook(samples, meta, output)
    print(f"Processed {len(samples)} samples.")
    print(f"DF overrides loaded: {len(df_overrides) // 2 if df_overrides else 0}")
    print(f"Output written to: {output}")


if __name__ == "__main__":
    main()
