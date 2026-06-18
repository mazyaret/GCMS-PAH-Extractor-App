#!/usr/bin/env python3
"""
GC/MS PAH Extractor - University of Iowa format v3.2

Purpose:
- Reads the University of Iowa GC/MS output format where:
  * The 16 PAHs are reported in one workbook with one worksheet per PAH.
  * Surrogates are reported in a second workbook with one worksheet per surrogate.
  * In each analyte worksheet, "Calc. Conc." is the result column, commonly column G.
  * Sample rows are identified by Type == "Sample"; calibration and QC rows are ignored.
- Reads the sample submission form for dilution factor and optional normalization inputs.
- Exports one processed workbook with raw/corrected concentrations, normalized values,
  totals, LMW/HMW summaries, TEQ summaries, and surrogate recovery.

Notes:
- UIowa Calc. Conc. values are treated as ng/mL because the calibration files and submission
  form use ng/mL. Output Corrected Conc. columns are reported in µg/mL by dividing by 1000 after DF correction.
- Excel table objects are intentionally not created to avoid Excel repair warnings.
"""

from __future__ import annotations

import argparse
import datetime
import math
import re
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule


TARGET_16_PAHS = [
    "Naphthalene",
    "Acenaphthylene",
    "Acenaphthene",
    "Fluorene",
    "Phenanthrene",
    "Anthracene",
    "Fluoranthene",
    "Pyrene",
    "Benz[a]anthracene",
    "Chrysene",
    "Benzo[b]fluoranthene",
    "Benzo[k]fluoranthene",
    "Benzo[a]pyrene",
    "Indeno[1,2,3-cd]pyrene",
    "Dibenz[a,h]anthracene",
    "Benzo[g,h,i]perylene",
]

LMW_PAHS = ["Naphthalene", "Acenaphthylene", "Acenaphthene", "Fluorene", "Phenanthrene", "Anthracene"]
HMW_PAHS = [a for a in TARGET_16_PAHS if a not in LMW_PAHS]

SURROGATES_UIOWA = [
    "Phenol-d6",
    "Nitrobenzene-d5",
    "2-Fluorobiphenyl",
    "2,4,6-Tribromophenol",
    "p-Terphenyl-d14",
]

# Recovery-correction mapping for the UIowa format.
# UIowa surrogate file has fewer PAH-like surrogates than the ISU file.
# Therefore LMW PAHs use 2-Fluorobiphenyl and HMW PAHs use p-Terphenyl-d14.
RECOVERY_SURROGATE_MAP_UIOWA = {
    "Naphthalene": "2-Fluorobiphenyl",
    "Acenaphthylene": "2-Fluorobiphenyl",
    "Acenaphthene": "2-Fluorobiphenyl",
    "Fluorene": "2-Fluorobiphenyl",
    "Phenanthrene": "2-Fluorobiphenyl",
    "Anthracene": "2-Fluorobiphenyl",
    "Fluoranthene": "p-Terphenyl-d14",
    "Pyrene": "p-Terphenyl-d14",
    "Benz[a]anthracene": "p-Terphenyl-d14",
    "Chrysene": "p-Terphenyl-d14",
    "Benzo[b]fluoranthene": "p-Terphenyl-d14",
    "Benzo[k]fluoranthene": "p-Terphenyl-d14",
    "Benzo[a]pyrene": "p-Terphenyl-d14",
    "Indeno[1,2,3-cd]pyrene": "p-Terphenyl-d14",
    "Dibenz[a,h]anthracene": "p-Terphenyl-d14",
    "Benzo[g,h,i]perylene": "p-Terphenyl-d14",
}


TEF_FACTORS = {
    "Naphthalene": 0.001,
    "Acenaphthylene": 0.001,
    "Acenaphthene": 0.001,
    "Fluorene": 0.001,
    "Phenanthrene": 0.001,
    "Anthracene": 0.01,
    "Fluoranthene": 0.001,
    "Pyrene": 0.001,
    "Benz[a]anthracene": 0.1,
    "Chrysene": 0.01,
    "Benzo[b]fluoranthene": 0.1,
    "Benzo[k]fluoranthene": 0.1,
    "Benzo[a]pyrene": 1,
    "Indeno[1,2,3-cd]pyrene": 0.1,
    "Dibenz[a,h]anthracene": 1,
    "Benzo[g,h,i]perylene": 0.01,
}

SHEET_ALIASES = {
    "naphthalene": "Naphthalene",
    "acenaphthalene": "Acenaphthylene",
    "acenaphthylene": "Acenaphthylene",
    "acenaphthene": "Acenaphthene",
    "fluorene": "Fluorene",
    "phenanthrene": "Phenanthrene",
    "anthracene": "Anthracene",
    "fluoranthene": "Fluoranthene",
    "pyrene": "Pyrene",
    "benz a anthracene": "Benz[a]anthracene",
    "benz[a]anthracene": "Benz[a]anthracene",
    "benzo a anthracene": "Benz[a]anthracene",
    "chrysene": "Chrysene",
    "benzo b fluoranthrene": "Benzo[b]fluoranthene",
    "benzo b fluoranthene": "Benzo[b]fluoranthene",
    "benzo[b]fluoranthrene": "Benzo[b]fluoranthene",
    "benzo[b]fluoranthene": "Benzo[b]fluoranthene",
    "benzo k fluoranthrene": "Benzo[k]fluoranthene",
    "benzo k fluoranthene": "Benzo[k]fluoranthene",
    "benzo[k]fluoranthrene": "Benzo[k]fluoranthene",
    "benzo[k]fluoranthene": "Benzo[k]fluoranthene",
    "benzo a pyrene": "Benzo[a]pyrene",
    "benzo[a]pyrene": "Benzo[a]pyrene",
    "indeno 1 2 3 cd pyrene": "Indeno[1,2,3-cd]pyrene",
    "indeno[1,2,3-cd]pyrene": "Indeno[1,2,3-cd]pyrene",
    "benzo g h i prylene": "Benzo[g,h,i]perylene",
    "benzo g h i perylene": "Benzo[g,h,i]perylene",
    "benzo[g,h,i]prylene": "Benzo[g,h,i]perylene",
    "benzo[g,h,i]perylene": "Benzo[g,h,i]perylene",
    "dibenz a h anthracene": "Dibenz[a,h]anthracene",
    "dibenz[a,h]anthracene": "Dibenz[a,h]anthracene",
    "dibenzo a h anthracene": "Dibenz[a,h]anthracene",
    "dibenzo[a,h]anthracene": "Dibenz[a,h]anthracene",
    # surrogates
    "phenol d6": "Phenol-d6",
    "phenol-d6": "Phenol-d6",
    "nitrobenzene d5": "Nitrobenzene-d5",
    "nitrobenzene-d5": "Nitrobenzene-d5",
    "2 fluorobiphenyl": "2-Fluorobiphenyl",
    "2-fluorobiphenyl": "2-Fluorobiphenyl",
    "2 4 6 tribromophenol": "2,4,6-Tribromophenol",
    "2,4,6-tribromophenol": "2,4,6-Tribromophenol",
    "p terphenyl d14": "p-Terphenyl-d14",
    "p-terphenyl-d14": "p-Terphenyl-d14",
    "terphenyl d14": "p-Terphenyl-d14",
    "terphenyl-d14": "p-Terphenyl-d14",
}


def norm_name(value: Any) -> str:
    text = str(value or "").strip()
    text = re.sub(r"\s+", " ", text)
    return text.replace("–", "-").replace("—", "-").lower()


def clean_key(value: Any) -> str:
    text = norm_name(value)
    text = re.sub(r"[\[\]\(\),]+", " ", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def standardize_sheet_analyte(sheet_name: Any) -> Optional[str]:
    key = clean_key(sheet_name)
    return SHEET_ALIASES.get(key)


def to_float_or_none(value: Any) -> Optional[float]:
    if value in (None, ""):
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, str):
        s = value.strip()
        if not s or s.lower() in {"none", "nan", "na", "n/a"}:
            return None
        # Accept decimal dot or decimal comma, and tolerate units such as "112.5 cm2".
        if "," in s and "." in s:
            s = s.replace(",", "")
        elif "," in s and "." not in s:
            s = s.replace(",", ".")
        m = re.search(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?", s)
        if m:
            try:
                return float(m.group(0))
            except Exception:
                return None
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def positive_or_none(value: Any) -> Optional[float]:
    v = to_float_or_none(value)
    if v is None or v <= 0:
        return None
    return v


def parse_df(value: Any) -> Tuple[float, str]:
    if value in (None, ""):
        return 1.0, "Default 1"
    if isinstance(value, (int, float)) and not (isinstance(value, float) and math.isnan(value)):
        return float(value), "Submission form"
    text = str(value).strip()
    if text.lower() in {"no", "none", "n/a", "na", "not diluted", "no dilution", "undiluted"}:
        return 1.0, "Submission form: no dilution"
    # e.g., 8X, x8, 1:8, diluted 8
    m = re.search(r"(?i)(\d+(?:\.\d+)?)\s*[x×]\b", text)
    if not m:
        m = re.search(r"(?i)\bx\s*(\d+(?:\.\d+)?)\b", text)
    if not m:
        m = re.search(r"(?i)\b1\s*[:/]\s*(\d+(?:\.\d+)?)\b", text)
    if not m:
        m = re.search(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?", text.replace(",", ""))
    if m:
        try:
            return float(m.group(1) if m.lastindex else m.group(0)), "Submission form"
        except Exception:
            pass
    return 1.0, "Default 1; unparsed dilution text"


def strip_datafile_extension(value: Any) -> str:
    text = str(value or "").strip()
    return re.sub(r"(?i)\.d$", "", text).strip()


def derive_sample_name_from_data_file(value: Any) -> str:
    text = strip_datafile_extension(value)
    # UIowa files often look like "1_P_047.D"; submission sample name is "P_047".
    text = re.sub(r"^\d+\s*[_-]\s*", "", text)
    return text.strip()


def match_keys(value: Any) -> List[str]:
    candidates = []
    raw = str(value or "").strip()
    if raw:
        candidates.append(raw)
        candidates.append(strip_datafile_extension(raw))
        candidates.append(derive_sample_name_from_data_file(raw))
    keys = []
    for c in candidates:
        if c:
            keys.extend([c, norm_name(c), clean_key(c)])
    seen = set()
    out = []
    for k in keys:
        if k and k not in seen:
            seen.add(k)
            out.append(k)
    return out


def _find_header_row_and_columns(ws) -> Tuple[int, Dict[str, Optional[int]]]:
    """Find the real submission-table header row.

    The UIowa template has a long instruction paragraph above the table that can contain
    words like "sample" and "name"; this function scores rows and chooses the row with
    actual compact column headers.
    """
    fields = {
        "sample_name": None,
        "dilution_factor": None,
        "extract_volume_ml": None,
        "fabric_size_cm2": None,
        "fabric_mass_g": None,
        "fuel_mass_loss_g": None,
        "surrogate_spike_ug_ml": None,
        "surrogate_spike_ng_ml": None,
    }
    best = None
    best_score = -1
    for r in range(1, min(ws.max_row, 100) + 1):
        found = fields.copy()
        sample_name_col = None
        tube_label_col = None
        score = 0
        for c in range(1, ws.max_column + 1):
            raw = ws.cell(r, c).value
            label = clean_key(raw)
            if not label:
                continue
            # Avoid treating long instruction sentences as headers.
            if len(label) > 60:
                continue
            if label in {"sample name", "sample"} or ("sample" in label and "name" in label and len(label) <= 30):
                sample_name_col = c
                score += 3
            if "tube" in label and "label" in label and len(label) <= 35:
                tube_label_col = c
                score += 1
            if label in {"df", "dilution factor"} or "dilution factor" in label or label == "dilution":
                found["dilution_factor"] = c
                score += 2
            if ("sample" in label and "volume" in label) or ("extract" in label and "volume" in label) or label in {"volume", "extract volume ml", "sample mass or volume ml"}:
                found["extract_volume_ml"] = c
                score += 1
            if ("fabric" in label and "size" in label) or "cm2" in label or "cm 2" in label:
                found["fabric_size_cm2"] = c
                score += 1
            if "fabric" in label and "mass" in label:
                found["fabric_mass_g"] = c
                score += 1
            if "fuel" in label and "mass" in label and "loss" in label:
                found["fuel_mass_loss_g"] = c
                score += 1
            # Surrogate spike/concentration column. Accept both explicit labels
            # like "Surrogate Spike (ug/mL)" and template labels like "Surrogate (ug/ml)".
            if (
                ("surrogate" in label and ("spike" in label or "conc" in label or "concentration" in label))
                or label in {"surrogate", "surrogate ug ml", "surrogate ug per ml"}
                or ("surrogate" in label and ("ug" in label or "microg" in label or "µg" in label) and "ml" in label)
            ):
                if "ng" in label:
                    found["surrogate_spike_ng_ml"] = c
                else:
                    found["surrogate_spike_ug_ml"] = c
                score += 1
        found["sample_name"] = sample_name_col or tube_label_col
        if found["sample_name"] and score > best_score:
            best = (r, found)
            best_score = score
    if best and best_score >= 2:
        return best
    raise ValueError("Could not find a sample-name column in the submission form.")


def load_submission_metadata(submission_xlsx: Optional[Path]) -> Dict[str, Dict[str, Any]]:
    if not submission_xlsx:
        return {}
    wb = load_workbook(submission_xlsx, data_only=True)
    # Prefer a sheet with "sample" in the name
    ws = None
    for s in wb.worksheets:
        if "sample" in norm_name(s.title):
            ws = s
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]

    header_row, cols = _find_header_row_and_columns(ws)
    metadata: Dict[str, Dict[str, Any]] = {}

    for r in range(header_row + 1, ws.max_row + 1):
        sample = ws.cell(r, cols["sample_name"]).value if cols.get("sample_name") else None
        if sample in (None, ""):
            continue
        sample_name = str(sample).strip()
        if not sample_name:
            continue
        raw_df = ws.cell(r, cols["dilution_factor"]).value if cols.get("dilution_factor") else None
        df, df_source = parse_df(raw_df)
        rec = {
            "sample_name": sample_name,
            "submission_row": r,
            "dilution_factor": df,
            "df_source": df_source,
            "extract_volume_ml": positive_or_none(ws.cell(r, cols["extract_volume_ml"]).value) if cols.get("extract_volume_ml") else 1.0,
            "fabric_size_cm2": positive_or_none(ws.cell(r, cols["fabric_size_cm2"]).value) if cols.get("fabric_size_cm2") else None,
            "fabric_mass_g": positive_or_none(ws.cell(r, cols["fabric_mass_g"]).value) if cols.get("fabric_mass_g") else None,
            "fuel_mass_loss_g": positive_or_none(ws.cell(r, cols["fuel_mass_loss_g"]).value) if cols.get("fuel_mass_loss_g") else None,
            "surrogate_spike_ug_ml": (
                (positive_or_none(ws.cell(r, cols["surrogate_spike_ug_ml"]).value) if cols.get("surrogate_spike_ug_ml") else None)
                if cols.get("surrogate_spike_ug_ml")
                else ((positive_or_none(ws.cell(r, cols["surrogate_spike_ng_ml"]).value) / 1000.0) if cols.get("surrogate_spike_ng_ml") and positive_or_none(ws.cell(r, cols["surrogate_spike_ng_ml"]).value) is not None else None)
            ),
            "parameter_source": "Uploaded UIowa submission file",
        }
        # Default extract volume to 1 mL if missing.
        if rec["extract_volume_ml"] is None:
            rec["extract_volume_ml"] = 1.0
        # Default UIowa surrogate spike to 0.05 µg/mL if missing, but keep source visible.
        if rec["surrogate_spike_ug_ml"] is None:
            rec["surrogate_spike_ug_ml"] = 0.05

        for key in match_keys(sample_name):
            metadata[key] = rec
    return metadata


def metadata_from_rows(rows: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    metadata: Dict[str, Dict[str, Any]] = {}
    for idx, row in enumerate(rows):
        sample_name = str(row.get("Sample Name") or row.get("sample_name") or "").strip()
        if not sample_name:
            continue
        rec = {
            "sample_name": sample_name,
            "submission_row": None,
            "dilution_factor": positive_or_none(row.get("Dilution Factor")) or 1.0,
            "df_source": "Manual in web app",
            "extract_volume_ml": positive_or_none(row.get("Extract Volume (mL)")) or 1.0,
            "fabric_size_cm2": positive_or_none(row.get("Fabric Size (cm^2)")),
            "fabric_mass_g": positive_or_none(row.get("Fabric Mass (g)")),
            "fuel_mass_loss_g": positive_or_none(row.get("Fuel Mass Loss (g)")),
            "surrogate_spike_ug_ml": positive_or_none(row.get("Surrogate Spike (µg/mL)")) or 0.05,
            "parameter_source": "Manual/reviewed in web app",
        }
        for key in match_keys(sample_name):
            metadata[key] = rec
    return metadata


def find_col_by_header(ws, header_name: str, header_row: int = 2) -> Optional[int]:
    target = clean_key(header_name)
    for c in range(1, ws.max_column + 1):
        if clean_key(ws.cell(header_row, c).value) == target:
            return c
    return None


def parse_result_workbook(result_xlsx: Path, allowed_analytes: List[str]) -> Tuple[Dict[str, Dict[str, float]], Dict[str, Any]]:
    wb = load_workbook(result_xlsx, data_only=True)
    values_by_sample: Dict[str, Dict[str, float]] = {}
    detected_sheets = {}
    warnings = []
    allowed_set = set(allowed_analytes)

    for ws in wb.worksheets:
        analyte = standardize_sheet_analyte(ws.title)
        if not analyte or analyte not in allowed_set:
            continue
        data_file_col = find_col_by_header(ws, "Data File", 2) or 1
        type_col = find_col_by_header(ws, "Type", 2) or 2
        calc_col = find_col_by_header(ws, "Calc. Conc.", 2) or 7
        detected_sheets[analyte] = ws.title

        for r in range(3, ws.max_row + 1):
            row_type = ws.cell(r, type_col).value
            if norm_name(row_type) != "sample":
                continue
            data_file = ws.cell(r, data_file_col).value
            if data_file in (None, ""):
                continue
            sample_name = derive_sample_name_from_data_file(data_file)
            raw = to_float_or_none(ws.cell(r, calc_col).value)
            if sample_name not in values_by_sample:
                values_by_sample[sample_name] = {"_data_file": str(data_file), "_first_row": r}
            values_by_sample[sample_name][analyte] = raw
    missing = [a for a in allowed_analytes if a not in detected_sheets]
    if missing:
        warnings.append(f"Missing expected analyte sheets: {missing}")
    meta = {
        "input_file": str(result_xlsx),
        "detected_sheets": detected_sheets,
        "missing_sheets": missing,
        "warnings": warnings,
    }
    return values_by_sample, meta


def get_metadata_for_sample(sample_name: str, data_file: Any, sample_metadata: Dict[str, Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    if not sample_metadata:
        return None
    for key in match_keys(sample_name) + match_keys(data_file):
        if key in sample_metadata:
            return sample_metadata[key]
    return None


def extract_uiowa(target_xlsx: Path, surrogate_xlsx: Path, sample_metadata: Optional[Dict[str, Dict[str, Any]]] = None):
    sample_metadata = sample_metadata or {}
    target_values, target_meta = parse_result_workbook(target_xlsx, TARGET_16_PAHS)
    surrogate_values, surrogate_meta = parse_result_workbook(surrogate_xlsx, SURROGATES_UIOWA)

    all_sample_names = []
    # Use submission order if available.
    seen_order = set()
    for rec in sample_metadata.values():
        nm = rec.get("sample_name")
        if nm and nm not in seen_order:
            all_sample_names.append(nm)
            seen_order.add(nm)
    for d in [target_values, surrogate_values]:
        for nm in d.keys():
            if nm not in seen_order:
                all_sample_names.append(nm)
                seen_order.add(nm)

    samples = []
    for nm in all_sample_names:
        data_file = None
        if nm in target_values:
            data_file = target_values[nm].get("_data_file")
        elif nm in surrogate_values:
            data_file = surrogate_values[nm].get("_data_file")
        md = get_metadata_for_sample(nm, data_file, sample_metadata) or {}
        display_name = md.get("sample_name") or nm

        df = float(md.get("dilution_factor") or 1.0)
        extract_volume_ml = md.get("extract_volume_ml") if md.get("extract_volume_ml") not in ("", None) else 1.0
        rec = {
            "sample_name": display_name,
            "uiowa_sample_key": nm,
            "data_file": data_file,
            "dilution_factor": df,
            "df_source": md.get("df_source", "Default 1"),
            "extract_volume_ml": positive_or_none(extract_volume_ml) or 1.0,
            "fabric_size_cm2": positive_or_none(md.get("fabric_size_cm2")),
            "fabric_mass_g": positive_or_none(md.get("fabric_mass_g")),
            "fuel_mass_loss_g": positive_or_none(md.get("fuel_mass_loss_g")),
            "surrogate_spike_ug_ml": positive_or_none(md.get("surrogate_spike_ug_ml")) or 0.05,
            "parameter_source": md.get("parameter_source", "Default/blank"),
            "target_values": {},
            "surrogate_values": {},
        }
        for analyte in TARGET_16_PAHS:
            rec["target_values"][analyte] = target_values.get(nm, {}).get(analyte)
        for surr in SURROGATES_UIOWA:
            rec["surrogate_values"][surr] = surrogate_values.get(nm, {}).get(surr)
        samples.append(rec)

    meta = {
        "format": "University of Iowa per-analyte worksheet format",
        "target_file": str(target_xlsx),
        "surrogate_file": str(surrogate_xlsx),
        "target_meta": target_meta,
        "surrogate_meta": surrogate_meta,
        "sample_count": len(samples),
        "created": datetime.datetime.now().isoformat(timespec="seconds"),
        "raw_calc_conc_unit": "ng/mL",
        "corrected_conc_unit": "µg/mL",
        "normalization_note": "Mass in ug = Corrected Conc. (µg/mL) * Extract Volume (mL)",
    }
    return samples, meta


def calc_norms(corrected_ug_ml: Optional[float], sample: Dict[str, Any]) -> Dict[str, Optional[float]]:
    if not isinstance(corrected_ug_ml, (int, float)):
        return {"mass_ug": None, "surface_load": None, "ug_g_fabric": None, "ug_g_fuel_loss": None}
    mass_ug = corrected_ug_ml * (sample.get("extract_volume_ml") or 1.0)
    fabric_size = sample.get("fabric_size_cm2")
    fabric_mass = sample.get("fabric_mass_g")
    fuel_loss = sample.get("fuel_mass_loss_g")
    return {
        "mass_ug": mass_ug,
        "surface_load": mass_ug * 100.0 / fabric_size if fabric_size else None,
        "ug_g_fabric": mass_ug / fabric_mass if fabric_mass else None,
        "ug_g_fuel_loss": mass_ug / fuel_loss if fuel_loss else None,
    }


def _sum_values(values: List[Optional[float]]) -> Optional[float]:
    nums = [v for v in values if isinstance(v, (int, float))]
    return sum(nums) if nums else None


def _summary_for_group(sample: Dict[str, Any], analytes: List[str], teq: bool = False) -> Dict[str, Optional[float]]:
    corrected_vals = []
    surface_vals = []
    fabric_vals = []
    fuel_vals = []
    for a in analytes:
        raw = sample["target_values"].get(a)
        if not isinstance(raw, (int, float)):
            continue
        corr = raw * sample["dilution_factor"] / 1000.0
        factor = TEF_FACTORS.get(a, 1.0) if teq else 1.0
        norms = calc_norms(corr, sample)
        corrected_vals.append(corr * factor)
        surface_vals.append(norms["surface_load"] * factor if norms["surface_load"] is not None else None)
        fabric_vals.append(norms["ug_g_fabric"] * factor if norms["ug_g_fabric"] is not None else None)
        fuel_vals.append(norms["ug_g_fuel_loss"] * factor if norms["ug_g_fuel_loss"] is not None else None)
    return {
        "corrected_ug_ml": _sum_values(corrected_vals),
        "surface_load": _sum_values(surface_vals),
        "ug_g_fabric": _sum_values(fabric_vals),
        "ug_g_fuel_loss": _sum_values(fuel_vals),
    }



def corrected_target_ug_ml(sample: Dict[str, Any], analyte: str) -> Optional[float]:
    raw = sample.get("target_values", {}).get(analyte)
    if isinstance(raw, (int, float)):
        return raw * sample.get("dilution_factor", 1.0) / 1000.0
    return None


def corrected_surrogate_ug_ml(sample: Dict[str, Any], surrogate: str) -> Optional[float]:
    raw = sample.get("surrogate_values", {}).get(surrogate)
    if isinstance(raw, (int, float)):
        return raw * sample.get("dilution_factor", 1.0) / 1000.0
    return None


def surrogate_recovery_pct_uiowa(sample: Dict[str, Any], surrogate: str) -> Optional[float]:
    corrected = corrected_surrogate_ug_ml(sample, surrogate)
    spike = sample.get("surrogate_spike_ug_ml") or 0.05
    if isinstance(corrected, (int, float)) and spike:
        return corrected / spike * 100.0
    return None


def recovery_corrected_conc_uiowa(sample: Dict[str, Any], analyte: str) -> Optional[float]:
    corrected = corrected_target_ug_ml(sample, analyte)
    surrogate = RECOVERY_SURROGATE_MAP_UIOWA.get(analyte)
    recovery = surrogate_recovery_pct_uiowa(sample, surrogate) if surrogate else None
    if isinstance(corrected, (int, float)) and isinstance(recovery, (int, float)) and recovery > 0:
        return corrected * 100.0 / recovery
    return None


def recovery_correction_factor_uiowa(sample: Dict[str, Any], analyte: str) -> Optional[float]:
    surrogate = RECOVERY_SURROGATE_MAP_UIOWA.get(analyte)
    recovery = surrogate_recovery_pct_uiowa(sample, surrogate) if surrogate else None
    if isinstance(recovery, (int, float)) and recovery > 0:
        return 100.0 / recovery
    return None


def _summary_for_group_recovery(sample: Dict[str, Any], analytes: List[str], teq: bool = False) -> Dict[str, Optional[float]]:
    corrected_vals = []
    surface_vals = []
    fabric_vals = []
    fuel_vals = []
    for a in analytes:
        corr = recovery_corrected_conc_uiowa(sample, a)
        if not isinstance(corr, (int, float)):
            continue
        factor = TEF_FACTORS.get(a, 1.0) if teq else 1.0
        norms = calc_norms(corr, sample)
        corrected_vals.append(corr * factor)
        surface_vals.append(norms["surface_load"] * factor if norms["surface_load"] is not None else None)
        fabric_vals.append(norms["ug_g_fabric"] * factor if norms["ug_g_fabric"] is not None else None)
        fuel_vals.append(norms["ug_g_fuel_loss"] * factor if norms["ug_g_fuel_loss"] is not None else None)
    return {
        "corrected_ug_ml": _sum_values(corrected_vals),
        "surface_load": _sum_values(surface_vals),
        "ug_g_fabric": _sum_values(fabric_vals),
        "ug_g_fuel_loss": _sum_values(fuel_vals),
    }


def _style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True, color="1F2937")
    thin = Side(style="thin", color="D9E2EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
    for col in range(1, ws.max_column + 1):
        max_len = 8
        for row in range(1, min(ws.max_row, 120) + 1):
            value = ws.cell(row, col).value
            max_len = max(max_len, min(len(str(value)) if value is not None else 0, 45))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 45)
    ws.freeze_panes = "A2"


def write_processed_workbook_uiowa(samples: List[Dict[str, Any]], meta: Dict[str, Any], output_path: Path) -> None:
    wb = Workbook()
    ws_info = wb.active
    ws_info.title = "Instructions"
    ws_samples = wb.create_sheet("Sample_Info")
    ws_wide = wb.create_sheet("Wide_Data")
    ws_summary = wb.create_sheet("Summary_Metrics")
    ws_long = wb.create_sheet("Long_Data")
    ws_qc = wb.create_sheet("Recovery_QC")
    ws_config = wb.create_sheet("Config_Analytes")
    ws_run = wb.create_sheet("Run_Info")

    ws_info.append(["University of Iowa GC/MS PAH Extraction Output v5"])
    ws_info.append(["Raw Calc. Conc. values from the UIowa files are treated as ng/mL."])
    ws_info.append(["Corrected Conc. (µg/mL) = Raw Calc. Conc. (ng/mL) × Dilution Factor ÷ 1000"])
    ws_info.append(["Surrogate recovery % = Corrected surrogate concentration (µg/mL) ÷ Surrogate Spike (µg/mL) × 100"])
    ws_info.append(["Recovery-Corrected Conc. = Corrected Conc. × 100 ÷ Assigned Surrogate Recovery %"])
    ws_info.append(["Recovery-corrected columns are added in addition to, not instead of, the original DF-corrected columns."])
    ws_info.append(["Mass (ug) = concentration (µg/mL) × Extract Volume (mL)"])
    ws_info.append(["Surface load (ug/100 cm^2) = Mass (ug) × 100 / Fabric Size (cm^2)"])
    ws_info.append(["ug/g fabric = Mass (ug) / Fabric Mass (g)"])
    ws_info.append(["ug/g fuel mass loss = Mass (ug) / Fuel Mass Loss (g)"])

    ws_samples.append([
        "Sample Name", "UIowa Sample Key", "Data File", "Dilution Factor", "DF Source",
        "Extract Volume (mL)", "Fabric Size (cm^2)", "Fabric Mass (g)", "Fuel Mass Loss (g)",
        "Surrogate Spike (µg/mL)", "Parameter Source",
    ])
    for s in samples:
        ws_samples.append([
            s["sample_name"], s["uiowa_sample_key"], s["data_file"], s["dilution_factor"], s["df_source"],
            s["extract_volume_ml"], s["fabric_size_cm2"], s["fabric_mass_g"], s["fuel_mass_loss_g"],
            s["surrogate_spike_ug_ml"], s["parameter_source"],
        ])

    wide_headers = [
        "Sample Name", "UIowa Sample Key", "Data File", "Dilution Factor", "DF Source",
        "Extract Volume (mL)", "Fabric Size (cm^2)", "Fabric Mass (g)", "Fuel Mass Loss (g)",
        "Surrogate Spike (µg/mL)", "Parameter Source",
    ]
    for analyte in TARGET_16_PAHS:
        wide_headers.extend([
            f"{analyte} Raw Calc. Conc. (ng/mL)",
            f"{analyte} Corrected Conc. (µg/mL)",
            f"{analyte} surface load (ug/100 cm^2)",
            f"{analyte} ug/g fabric",
            f"{analyte} ug/g mass loss of fuel",
            f"{analyte} Assigned Recovery Surrogate",
            f"{analyte} Surrogate Recovery %",
            f"{analyte} Recovery Correction Factor",
            f"{analyte} Recovery-Corrected Conc. (µg/mL)",
            f"{analyte} Recovery-Corrected surface load (ug/100 cm^2)",
            f"{analyte} Recovery-Corrected ug/g fabric",
            f"{analyte} Recovery-Corrected ug/g mass loss of fuel",
        ])
    for label in ["Total PAHs", "LMW PAHs", "HMW PAHs"]:
        wide_headers.extend([f"{label} Corrected Conc. (µg/mL)", f"{label} surface load (ug/100 cm^2)", f"{label} ug/g fabric", f"{label} ug/g mass loss of fuel"])
    for label in ["Total PAHs TEQ", "LMW PAHs TEQ", "HMW PAHs TEQ"]:
        wide_headers.extend([f"{label} Corrected Conc. (µg/mL)", f"{label} surface load (ug/100 cm^2)", f"{label} ug/g fabric", f"{label} ug/g mass loss of fuel"])
    for label in ["Total PAHs", "LMW PAHs", "HMW PAHs"]:
        wide_headers.extend([f"{label} Recovery-Corrected Conc. (µg/mL)", f"{label} Recovery-Corrected surface load (ug/100 cm^2)", f"{label} Recovery-Corrected ug/g fabric", f"{label} Recovery-Corrected ug/g mass loss of fuel"])
    for label in ["Total PAHs TEQ", "LMW PAHs TEQ", "HMW PAHs TEQ"]:
        wide_headers.extend([f"{label} Recovery-Corrected Conc. (µg/mL)", f"{label} Recovery-Corrected surface load (ug/100 cm^2)", f"{label} Recovery-Corrected ug/g fabric", f"{label} Recovery-Corrected ug/g mass loss of fuel"])
    for surr in SURROGATES_UIOWA:
        wide_headers.extend([f"{surr} Raw Calc. Conc. (ng/mL)", f"{surr} Corrected Conc. (µg/mL)", f"{surr} Recovery %"])
    ws_wide.append(wide_headers)

    for s in samples:
        row = [s["sample_name"], s["uiowa_sample_key"], s["data_file"], s["dilution_factor"], s["df_source"], s["extract_volume_ml"], s["fabric_size_cm2"], s["fabric_mass_g"], s["fuel_mass_loss_g"], s["surrogate_spike_ug_ml"], s["parameter_source"]]
        for analyte in TARGET_16_PAHS:
            raw = s["target_values"].get(analyte)
            corrected = corrected_target_ug_ml(s, analyte)
            norms = calc_norms(corrected, s)
            assigned = RECOVERY_SURROGATE_MAP_UIOWA.get(analyte)
            rec_pct = surrogate_recovery_pct_uiowa(s, assigned) if assigned else None
            rec_factor = recovery_correction_factor_uiowa(s, analyte)
            rec_corr = recovery_corrected_conc_uiowa(s, analyte)
            rec_norms = calc_norms(rec_corr, s)
            row.extend([raw, corrected, norms["surface_load"], norms["ug_g_fabric"], norms["ug_g_fuel_loss"], assigned, rec_pct, rec_factor, rec_corr, rec_norms["surface_load"], rec_norms["ug_g_fabric"], rec_norms["ug_g_fuel_loss"]])
        for analytes in [TARGET_16_PAHS, LMW_PAHS, HMW_PAHS]:
            summ = _summary_for_group(s, analytes, teq=False)
            row.extend([summ["corrected_ug_ml"], summ["surface_load"], summ["ug_g_fabric"], summ["ug_g_fuel_loss"]])
        for analytes in [TARGET_16_PAHS, LMW_PAHS, HMW_PAHS]:
            summ = _summary_for_group(s, analytes, teq=True)
            row.extend([summ["corrected_ug_ml"], summ["surface_load"], summ["ug_g_fabric"], summ["ug_g_fuel_loss"]])
        for analytes in [TARGET_16_PAHS, LMW_PAHS, HMW_PAHS]:
            summ = _summary_for_group_recovery(s, analytes, teq=False)
            row.extend([summ["corrected_ug_ml"], summ["surface_load"], summ["ug_g_fabric"], summ["ug_g_fuel_loss"]])
        for analytes in [TARGET_16_PAHS, LMW_PAHS, HMW_PAHS]:
            summ = _summary_for_group_recovery(s, analytes, teq=True)
            row.extend([summ["corrected_ug_ml"], summ["surface_load"], summ["ug_g_fabric"], summ["ug_g_fuel_loss"]])
        for surr in SURROGATES_UIOWA:
            raw = s["surrogate_values"].get(surr)
            corrected = corrected_surrogate_ug_ml(s, surr)
            spike = s.get("surrogate_spike_ug_ml") or 0.05
            recovery = corrected / spike * 100 if isinstance(corrected, (int, float)) and spike else None
            row.extend([raw, corrected, recovery])
        ws_wide.append(row)

    ws_summary.append([
        "Sample Name", "Category",
        "Corrected Conc. (µg/mL)", "surface load (ug/100 cm^2)", "ug/g fabric", "ug/g mass loss of fuel",
        "TEQ Corrected Conc. (µg/mL)", "TEQ surface load (ug/100 cm^2)", "TEQ ug/g fabric", "TEQ ug/g mass loss of fuel",
        "Recovery-Corrected Conc. (µg/mL)", "Recovery-Corrected surface load (ug/100 cm^2)", "Recovery-Corrected ug/g fabric", "Recovery-Corrected ug/g mass loss of fuel",
        "Recovery-Corrected TEQ Conc. (µg/mL)", "Recovery-Corrected TEQ surface load (ug/100 cm^2)", "Recovery-Corrected TEQ ug/g fabric", "Recovery-Corrected TEQ ug/g mass loss of fuel",
    ])
    for s in samples:
        for label, group in [("Total PAHs", TARGET_16_PAHS), ("LMW PAHs", LMW_PAHS), ("HMW PAHs", HMW_PAHS)]:
            m = _summary_for_group(s, group, teq=False)
            t = _summary_for_group(s, group, teq=True)
            r = _summary_for_group_recovery(s, group, teq=False)
            rt = _summary_for_group_recovery(s, group, teq=True)
            ws_summary.append([s["sample_name"], label, m["corrected_ug_ml"], m["surface_load"], m["ug_g_fabric"], m["ug_g_fuel_loss"], t["corrected_ug_ml"], t["surface_load"], t["ug_g_fabric"], t["ug_g_fuel_loss"], r["corrected_ug_ml"], r["surface_load"], r["ug_g_fabric"], r["ug_g_fuel_loss"], rt["corrected_ug_ml"], rt["surface_load"], rt["ug_g_fabric"], rt["ug_g_fuel_loss"]])

    ws_long.append([
        "Sample Name", "UIowa Sample Key", "Data File", "Dilution Factor", "DF Source", "Extract Volume (mL)", "Fabric Size (cm^2)", "Fabric Mass (g)", "Fuel Mass Loss (g)", "Analyte", "Role",
        "Raw Calc. Conc. (ng/mL)", "Corrected Conc. (µg/mL)", "Mass (ug)", "surface load (ug/100 cm^2)", "ug/g fabric", "ug/g mass loss of fuel",
        "Assigned Recovery Surrogate", "Surrogate Recovery %", "Recovery Correction Factor", "Recovery-Corrected Conc. (µg/mL)", "Recovery-Corrected Mass (ug)", "Recovery-Corrected surface load (ug/100 cm^2)", "Recovery-Corrected ug/g fabric", "Recovery-Corrected ug/g mass loss of fuel",
        "TEF", "TEQ Corrected Conc. (µg/mL)", "TEQ Recovery-Corrected Conc. (µg/mL)",
    ])
    for s in samples:
        for analyte in TARGET_16_PAHS:
            raw = s["target_values"].get(analyte)
            corr = corrected_target_ug_ml(s, analyte)
            norms = calc_norms(corr, s)
            assigned = RECOVERY_SURROGATE_MAP_UIOWA.get(analyte)
            rec_pct = surrogate_recovery_pct_uiowa(s, assigned) if assigned else None
            rec_factor = recovery_correction_factor_uiowa(s, analyte)
            rec_corr = recovery_corrected_conc_uiowa(s, analyte)
            rec_norms = calc_norms(rec_corr, s)
            tef = TEF_FACTORS.get(analyte)
            ws_long.append([s["sample_name"], s["uiowa_sample_key"], s["data_file"], s["dilution_factor"], s["df_source"], s["extract_volume_ml"], s["fabric_size_cm2"], s["fabric_mass_g"], s["fuel_mass_loss_g"], analyte, "Target", raw, corr, norms["mass_ug"], norms["surface_load"], norms["ug_g_fabric"], norms["ug_g_fuel_loss"], assigned, rec_pct, rec_factor, rec_corr, rec_norms["mass_ug"], rec_norms["surface_load"], rec_norms["ug_g_fabric"], rec_norms["ug_g_fuel_loss"], tef, corr * tef if isinstance(corr, (int, float)) else None, rec_corr * tef if isinstance(rec_corr, (int, float)) else None])

    ws_qc.append(["Sample Name", "Surrogate", "Raw Calc. Conc. (ng/mL)", "Corrected Conc. (µg/mL)", "Spike Conc. (µg/mL)", "Recovery %", "QC Flag"])
    for s in samples:
        for surr in SURROGATES_UIOWA:
            raw = s["surrogate_values"].get(surr)
            corrected = corrected_surrogate_ug_ml(s, surr)
            spike = s.get("surrogate_spike_ug_ml") or 0.05
            recovery = corrected / spike * 100 if isinstance(corrected, (int, float)) and spike else None
            flag = "OK" if isinstance(recovery, (int, float)) and 50 <= recovery <= 150 else "Check"
            ws_qc.append([s["sample_name"], surr, raw, corrected, spike, recovery, flag])

    ws_config.append(["Analyte", "Role", "Group", "TEF", "Assigned Recovery Surrogate", "Notes"])
    for a in TARGET_16_PAHS:
        ws_config.append([a, "Target", "LMW" if a in LMW_PAHS else "HMW", TEF_FACTORS.get(a), RECOVERY_SURROGATE_MAP_UIOWA.get(a), "UIowa PAH target"])
    for surr in SURROGATES_UIOWA:
        ws_config.append([surr, "Surrogate", "", None, None, "UIowa surrogate"])

    ws_run.append(["Field", "Value"])
    for k, v in meta.items():
        ws_run.append([k, str(v)])
    ws_run.append(["Recovery correction formula", "Recovery-Corrected Conc. = Corrected Conc. × 100 / Assigned Surrogate Recovery %"])
    ws_run.append(["Created", datetime.datetime.now().isoformat(timespec="seconds")])

    for ws in wb.worksheets:
        _style_sheet(ws)
    for ws in [ws_wide, ws_summary, ws_long, ws_qc]:
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = "0.0000"
    if ws_qc.max_row > 1:
        red_fill = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")
        ws_qc.conditional_formatting.add(f"G2:G{ws_qc.max_row}", CellIsRule(operator="equal", formula=['"Check"'], fill=red_fill))
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="Extract University of Iowa GC/MS PAH and surrogate files.")
    parser.add_argument("--submission", type=Path, required=True)
    parser.add_argument("--targets", type=Path, required=True)
    parser.add_argument("--surrogates", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    metadata = load_submission_metadata(args.submission)
    samples, meta = extract_uiowa(args.targets, args.surrogates, sample_metadata=metadata)
    write_processed_workbook_uiowa(samples, meta, args.output)
    print(f"Processed {len(samples)} UIowa samples.")
    print(f"Output written to: {args.output}")


if __name__ == "__main__":
    main()
